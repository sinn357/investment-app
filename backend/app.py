from flask import Flask, jsonify, request, make_response, Response, send_file
from flask_cors import CORS
import os
import functools
from dotenv import load_dotenv
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import requests
import csv
import time

# 통합 크롤러
from crawlers.unified_crawler import crawl_indicator, crawl_category
from crawlers.indicators_config import INDICATORS, CATEGORIES, get_all_enabled_indicators, get_indicator_config
from services.database_service import DatabaseService
from services.crawler_service import CrawlerService
from services.macro_cycle_service import MacroCycleService
from services.credit_cycle_service import CreditCycleService
from metadata.indicator_metadata import IndicatorMetadata
import threading
import time
import json

# Redis 임포트 시도 (Phase 3 캐싱)
try:
    import redis
    REDIS_AVAILABLE = True
except ImportError:
    print("Redis not available")
    REDIS_AVAILABLE = False

# PostgreSQL 서비스 임포트 시도
try:
    from services.postgres_database_service import PostgresDatabaseService
    POSTGRES_AVAILABLE = True
except ImportError as e:
    print(f"PostgreSQL not available: {e}")
    POSTGRES_AVAILABLE = False

load_dotenv()

app = Flask(__name__)
CORS(app,
     origins=["https://investment-app-rust-one.vercel.app", "http://localhost:3000"],
     methods=['GET', 'POST', 'PUT', 'DELETE', 'PATCH', 'OPTIONS'],
     allow_headers=['Content-Type', 'Authorization'],
     supports_credentials=True)

# ✅ Phase 3: Redis 클라이언트 초기화
redis_client = None
if REDIS_AVAILABLE:
    try:
        redis_url = os.getenv('REDIS_URL')
        if redis_url:
            redis_client = redis.from_url(redis_url, decode_responses=True)
            # 연결 테스트
            redis_client.ping()
            print("✅ Redis connected successfully")
        else:
            print("⚠️ REDIS_URL not set, Redis caching disabled")
    except Exception as e:
        print(f"⚠️ Redis connection failed: {e}")
        redis_client = None

# CORS preflight 핸들러 추가
@app.before_request
def handle_preflight():
    if request.method == "OPTIONS":
        response = Response()
        origin = request.headers.get('Origin')
        allowed_origins = ["https://investment-app-rust-one.vercel.app", "http://localhost:3000"]
        if origin in allowed_origins:
            response.headers.add("Access-Control-Allow-Origin", origin)
        response.headers.add("Access-Control-Allow-Headers", "Content-Type, Authorization")
        response.headers.add("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, PATCH, OPTIONS")
        response.headers.add("Access-Control-Allow-Credentials", "true")
        return response

# 모든 응답에 CORS 헤더 추가 (중복 방지를 위해 set 사용)
@app.after_request
def add_cors_headers(response):
    origin = request.headers.get('Origin')
    allowed_origins = ["https://investment-app-rust-one.vercel.app", "http://localhost:3000"]
    if origin in allowed_origins:
        # add 대신 direct assignment로 중복 방지
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
    return response

# 데이터베이스 서비스 초기화 (PostgreSQL 우선 사용)
try:
    database_url = os.getenv('DATABASE_URL')
    if database_url and database_url.startswith('postgres') and POSTGRES_AVAILABLE:
        print("Using PostgreSQL database...")
        db_service = PostgresDatabaseService(database_url)
    else:
        print("Using SQLite database...")
        db_service = DatabaseService()
except Exception as e:
    print(f"PostgreSQL connection failed, falling back to SQLite: {e}")
    db_service = DatabaseService()

# JWT 토큰 검증 데코레이터
def token_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        token = None

        # Authorization 헤더에서 토큰 추출
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(" ")[1]  # "Bearer <token>"
            except IndexError:
                return jsonify({'status': 'error', 'message': 'Invalid token format'}), 401

        if not token:
            return jsonify({'status': 'error', 'message': 'Token is missing'}), 401

        try:
            # JWT 토큰 검증
            current_user = db_service.verify_jwt_token(token)
            if current_user['status'] != 'success':
                return jsonify(current_user), 401
        except Exception as e:
            print(f"Token verification error: {e}")
            return jsonify({'status': 'error', 'message': 'Token is invalid'}), 401

        return f(current_user, *args, **kwargs)
    return decorated

# 보안 헤더 추가
@app.after_request
def after_request(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# 업데이트 상태 추적 (전역 변수)
update_status = {
    "is_updating": False,
    "progress": 0,
    "current_indicator": "",
    "completed_indicators": [],
    "failed_indicators": [],
    "total_indicators": 0,  # 전체 지표 개수
    "start_time": None
}

UPDATE_STATUS_REDIS_KEY = "indicators:update_status"

def load_update_status():
    """Redis에 저장된 업데이트 상태가 있으면 우선 사용 (멀티 인스턴스 대응)"""
    if redis_client:
        try:
            cached = redis_client.get(UPDATE_STATUS_REDIS_KEY)
            if cached:
                return json.loads(cached)
        except Exception as e:
            print(f"⚠️ Redis get update_status error: {e}")
    return update_status

def save_update_status():
    """업데이트 상태를 Redis에 저장 (TTL 1시간)"""
    if redis_client:
        try:
            redis_client.setex(
                UPDATE_STATUS_REDIS_KEY,
                3600,
                json.dumps(update_status)
            )
        except Exception as e:
            print(f"⚠️ Redis set update_status error: {e}")

# 인메모리 캐시 (지표 전체 조회용)
INDICATORS_CACHE = {"data": None, "timestamp": 0}
INDICATORS_CACHE_TTL = 300  # seconds
MAX_UPDATE_DURATION = 600  # seconds, 오래 걸리면 스테일 처리

def _missing_retail_sales(*_args, **_kwargs):
    return {"error": "Retail sales crawler not available"}


def _missing_retail_sales_yoy(*_args, **_kwargs):
    return {"error": "Retail sales YoY crawler not available"}


try:
    print("Testing retail sales import...")
    from crawlers.retail_sales import get_retail_sales_data
    print("Retail sales import successful")
except Exception as e:
    print(f"Retail sales import failed: {e}")
    try:
        from crawlers.archive_old_crawlers.retail_sales import get_retail_sales_data
        print("Retail sales import fallback successful")
    except Exception as fallback_error:
        print(f"Retail sales import fallback failed: {fallback_error}")
        get_retail_sales_data = _missing_retail_sales

try:
    from crawlers.retail_sales_yoy import get_retail_sales_yoy_data
    print("Retail sales YoY import successful")
except Exception as e:
    print(f"Retail sales YoY import failed: {e}")
    try:
        from crawlers.archive_old_crawlers.retail_sales_yoy import get_retail_sales_yoy_data
        print("Retail sales YoY import fallback successful")
    except Exception as fallback_error:
        print(f"Retail sales YoY import fallback failed: {fallback_error}")
        get_retail_sales_yoy_data = _missing_retail_sales_yoy

@app.route('/')
def health_check():
    """GitHub Actions용 헬스체크 엔드포인트 (루트 경로)"""
    try:
        # DB 연결 확인
        db_status = "connected" if db_service else "disconnected"

        # 간단한 DB 쿼리로 실제 연결 테스트
        if db_service:
            try:
                with db_service.get_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT 1")
                        cur.fetchone()
                db_status = "healthy"
            except Exception as db_error:
                db_status = f"error: {str(db_error)}"

        return jsonify({
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "database": db_status,
            "service": "investment-app-backend"
        })
    except Exception as e:
        return jsonify({
            "status": "unhealthy",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }), 500

@app.route('/api/health', methods=['GET'])
def api_health_check():
    """GitHub Actions용 헬스체크 엔드포인트 (/api/health)"""
    return health_check()

@app.route('/api/economic-indicators')
def get_economic_indicators():
    mock_data = [
        {
            "name": "한국은행 기준금리",
            "latestDate": "2025-09-15",
            "nextDate": "2025-10-15",
            "actual": 3.50,
            "forecast": 3.25,
            "previous": 3.25,
            "surprise": 0.25,
            "threshold": {"value": 4.0, "type": "warning"}
        },
        {
            "name": "소비자물가지수(CPI)",
            "latestDate": "2025-09-10",
            "nextDate": "2025-10-10",
            "actual": 2.8,
            "forecast": 2.5,
            "previous": 2.6,
            "surprise": 0.3,
            "threshold": {"value": 3.0, "type": "critical"}
        },
        {
            "name": "실업률",
            "latestDate": "2025-09-08",
            "nextDate": "2025-10-08",
            "actual": 3.2,
            "forecast": 3.1,
            "previous": 3.0,
            "surprise": 0.1,
            "threshold": None
        },
        {
            "name": "GDP 성장률",
            "latestDate": "2025-08-30",
            "nextDate": "2025-11-30",
            "actual": None,
            "forecast": 2.8,
            "previous": 2.6,
            "surprise": None,
            "threshold": None
        },
        {
            "name": "산업생산지수",
            "latestDate": "2025-09-12",
            "nextDate": "2025-10-12",
            "actual": 1.2,
            "forecast": 0.8,
            "previous": 0.5,
            "surprise": 0.4,
            "threshold": {"value": 2.0, "type": "warning"}
        }
    ]

    return jsonify({
        "status": "success",
        "data": mock_data,
        "timestamp": "2025-09-17T13:20:00Z"
    })

@app.route('/api/rawdata/latest')
def get_latest_rawdata():
    """Get latest economic data from investing.com crawling"""
    try:
        data = get_ism_manufacturing_pmi()

        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "ISM Manufacturing PMI"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/<indicator>')
def get_history_table(indicator):
    """Get full history table data for specific indicator"""
    try:
        # Map indicator names to URLs
        indicator_urls = {
            'ism-manufacturing': 'https://www.investing.com/economic-calendar/ism-manufacturing-pmi-173',
            'ism-non-manufacturing': 'https://www.investing.com/economic-calendar/ism-non-manufacturing-pmi-176',
            'sp-global-composite': 'https://www.investing.com/economic-calendar/s-p-global-composite-pmi-1492',
            'industrial-production': 'https://www.investing.com/economic-calendar/industrial-production-161',
            'industrial-production-1755': 'https://www.investing.com/economic-calendar/industrial-production-1755'
        }

        if indicator not in indicator_urls:
            return jsonify({
                "status": "error",
                "message": f"Indicator '{indicator}' not supported"
            }), 404

        url = indicator_urls[indicator]

        # Fetch and parse history table
        html = fetch_html(url)
        rows = parse_history_table(html)

        return jsonify({
            "status": "success",
            "indicator": indicator,
            "data": rows,
            "source": "investing.com",
            "total_rows": len(rows)
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Failed to fetch history table: {str(e)}"
        }), 500

@app.route('/api/rawdata/ism-non-manufacturing')
def get_ism_non_manufacturing_rawdata():
    """Get ISM Non-Manufacturing PMI raw data"""
    try:
        data = get_ism_non_manufacturing_pmi()

        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "ISM Non-Manufacturing PMI"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/rawdata/sp-global-composite')
def get_sp_global_composite_rawdata():
    """Get S&P Global Composite PMI raw data"""
    try:
        data = get_sp_global_composite_pmi()

        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "S&P Global Composite PMI"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/rawdata/industrial-production')
def get_industrial_production_rawdata():
    """Get Industrial Production raw data"""
    try:
        data = get_industrial_production()

        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Industrial Production"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/rawdata/industrial-production-1755')
def get_industrial_production_1755_rawdata():
    """Get Industrial Production (1755) raw data"""
    try:
        data = get_industrial_production_1755()

        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Industrial Production YoY"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/rawdata/retail-sales')
def get_retail_sales_rawdata():
    try:
        data = get_retail_sales_data()

        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Retail Sales MoM"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/retail-sales')
def get_retail_sales_history():
    try:
        url = "https://www.investing.com/economic-calendar/retail-sales-256"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            if history_data:
                return jsonify({
                    "status": "success",
                    "data": history_data,
                    "indicator": "Retail Sales MoM",
                    "source": "investing.com"
                })
        return jsonify({"status": "error", "message": "Failed to fetch retail sales history data"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route('/api/rawdata/retail-sales-yoy')
def get_retail_sales_yoy_rawdata():
    try:
        data = get_retail_sales_yoy_data()

        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Retail Sales YoY"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/retail-sales-yoy')
def get_retail_sales_yoy_history():
    try:
        url = "https://www.investing.com/economic-calendar/retail-sales-1878"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            if history_data:
                return jsonify({
                    "status": "success",
                    "data": history_data,
                    "indicator": "Retail Sales YoY",
                    "source": "investing.com"
                })
        return jsonify({"status": "error", "message": "Failed to fetch retail sales YoY history data"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route('/api/rawdata/gdp')
def get_gdp_rawdata():
    try:
        data = get_gdp_data()
        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500
        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "GDP QoQ"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/gdp')
def get_gdp_history():
    try:
        url = "https://www.investing.com/economic-calendar/gdp-375"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            if history_data:
                return jsonify({
                    "status": "success",
                    "data": history_data,
                    "indicator": "GDP QoQ",
                    "source": "investing.com"
                })
        return jsonify({"status": "error", "message": "Failed to fetch GDP history data"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route('/api/rawdata/cb-consumer-confidence')
def get_cb_consumer_confidence_rawdata():
    try:
        data = get_cb_consumer_confidence_data()
        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500
        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "CB Consumer Confidence"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/cb-consumer-confidence')
def get_cb_consumer_confidence_history():
    try:
        url = "https://www.investing.com/economic-calendar/cb-consumer-confidence-48"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            if history_data:
                return jsonify({
                    "status": "success",
                    "data": history_data,
                    "indicator": "CB Consumer Confidence",
                    "source": "investing.com"
                })
        return jsonify({"status": "error", "message": "Failed to fetch CB Consumer Confidence history data"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route('/api/rawdata/michigan-consumer-sentiment')
def get_michigan_consumer_sentiment_rawdata():
    try:
        data = get_michigan_consumer_sentiment_data()
        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500
        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Michigan Consumer Sentiment"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/michigan-consumer-sentiment')
def get_michigan_consumer_sentiment_history():
    try:
        url = "https://www.investing.com/economic-calendar/michigan-consumer-sentiment-320"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            if history_data:
                return jsonify({
                    "status": "success",
                    "data": history_data,
                    "indicator": "Michigan Consumer Sentiment",
                    "source": "investing.com"
                })
        return jsonify({"status": "error", "message": "Failed to fetch Michigan Consumer Sentiment history data"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

# ========== 고용지표 API 엔드포인트 ==========

@app.route('/api/rawdata/unemployment-rate')
def get_unemployment_rate_rawdata():
    """실업률 Raw Data API 엔드포인트"""
    try:
        data = get_unemployment_rate()
        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500
        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Unemployment Rate"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/rawdata/nonfarm-payrolls')
def get_nonfarm_payrolls_rawdata():
    """비농업 고용 Raw Data API 엔드포인트"""
    try:
        data = get_nonfarm_payrolls()
        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500
        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Nonfarm Payrolls"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/unemployment-rate')
def get_unemployment_rate_history():
    """실업률 History Table API 엔드포인트"""
    try:
        url = "https://www.investing.com/economic-calendar/unemployment-rate-300"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            if history_data:
                return jsonify({
                    "status": "success",
                    "data": history_data,
                    "indicator": "Unemployment Rate",
                    "source": "investing.com"
                })
        return jsonify({"status": "error", "message": "Failed to fetch Unemployment Rate history data"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route('/api/history-table/nonfarm-payrolls')
def get_nonfarm_payrolls_history():
    """비농업 고용 History Table API 엔드포인트"""
    try:
        url = "https://www.investing.com/economic-calendar/nonfarm-payrolls"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            if history_data:
                return jsonify({
                    "status": "success",
                    "data": history_data,
                    "indicator": "Nonfarm Payrolls",
                    "source": "investing.com"
                })
        return jsonify({"status": "error", "message": "Failed to fetch Nonfarm Payrolls history data"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

# ========== 새로운 고용지표 API 엔드포인트 ==========

@app.route('/api/rawdata/initial-jobless-claims')
def get_initial_jobless_claims_rawdata():
    """Initial Jobless Claims Raw Data API 엔드포인트"""
    try:
        data = get_initial_jobless_claims()
        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500
        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Initial Jobless Claims"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/initial-jobless-claims')
def get_initial_jobless_claims_history():
    """Initial Jobless Claims History Table API 엔드포인트"""
    try:
        url = "https://www.investing.com/economic-calendar/initial-jobless-claims-294"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            if history_data:
                return jsonify({
                    "status": "success",
                    "data": history_data,
                    "indicator": "Initial Jobless Claims",
                    "source": "investing.com"
                })
        return jsonify({"status": "error", "message": "Failed to fetch Initial Jobless Claims history data"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route('/api/rawdata/average-hourly-earnings')
def get_average_hourly_earnings_rawdata():
    """Average Hourly Earnings Raw Data API 엔드포인트"""
    try:
        data = get_average_hourly_earnings()
        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500
        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Average Hourly Earnings"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/average-hourly-earnings')
def get_average_hourly_earnings_history():
    """Average Hourly Earnings History Table API 엔드포인트"""
    try:
        url = "https://www.investing.com/economic-calendar/average-hourly-earnings-8"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            if history_data:
                return jsonify({
                    "status": "success",
                    "data": history_data,
                    "indicator": "Average Hourly Earnings",
                    "source": "investing.com"
                })
        return jsonify({"status": "error", "message": "Failed to fetch Average Hourly Earnings history data"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route('/api/rawdata/average-hourly-earnings-1777')
def get_average_hourly_earnings_1777_rawdata():
    """Average Hourly Earnings (1777) Raw Data API 엔드포인트"""
    try:
        data = get_average_hourly_earnings_1777()
        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500
        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Average Hourly Earnings (YoY)"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/average-hourly-earnings-1777')
def get_average_hourly_earnings_1777_history():
    """Average Hourly Earnings (1777) History Table API 엔드포인트"""
    try:
        url = "https://www.investing.com/economic-calendar/average-hourly-earnings-1777"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            if history_data:
                return jsonify({
                    "status": "success",
                    "data": history_data,
                    "indicator": "Average Hourly Earnings (YoY)",
                    "source": "investing.com"
                })
        return jsonify({"status": "error", "message": "Failed to fetch Average Hourly Earnings (1777) history data"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route('/api/rawdata/participation-rate')
def get_participation_rate_rawdata():
    """Participation Rate Raw Data API 엔드포인트"""
    try:
        data = get_participation_rate()
        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500
        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Participation Rate"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/participation-rate')
def get_participation_rate_history():
    """Participation Rate History Table API 엔드포인트"""
    try:
        url = "https://www.investing.com/economic-calendar/participation-rate-1581"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            if history_data:
                return jsonify({
                    "status": "success",
                    "data": history_data,
                    "indicator": "Participation Rate",
                    "source": "investing.com"
                })
        return jsonify({"status": "error", "message": "Failed to fetch Participation Rate history data"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

# ========== 새로운 데이터베이스 기반 API 엔드포인트 ==========

CATEGORY_LABELS = {
    "business": "경기",
    "employment": "고용",
    "interest": "금리",
    "trade": "무역",
    "inflation": "물가",
    "credit": "신용",
    "sentiment": "심리",
}

CATEGORY_SECTION_TITLES = {
    "business": [
        "현재 경기 상태",
        "속도 변화",
        "무엇이 경제를 이끄는지",
        "금리 방향에 미칠 영향",
        "자산시장 영향",
        "한 줄 요약",
    ],
    "employment": [
        "현재 고용 상태 (강함 / 둔화 / 위험 신호)",
        "해고 위험 여부",
        "임금 압력 수준",
        "소비 지속 가능성",
        "금리 방향에 미칠 영향",
        "자산시장 영향 (성장주 / 소비주 / 채권 / 달러)",
        "한 줄 요약",
    ],
    "inflation": [
        "인플레이션 방향 (상승 / 둔화 / 정체)",
        "속도 변화 (MoM 기준)",
        "구조적 물가 압력 존재 여부",
        "금리 인하 가능성에 미칠 영향",
        "자산시장 영향 (금리 / 성장주 / 원자재 / 달러)",
        "한 줄 요약",
    ],
    "interest": [
        "현재 정책 상태 (긴축 / 중립 / 완화)",
        "시장의 금리 기대 (인하 기대 / 유지 / 긴축 지속)",
        "성장 기대 해석 (10년물 기준)",
        "실질금리의 자산시장 영향",
        "달러 및 위험자산에 미칠 영향",
        "한 줄 요약",
    ],
    "trade": [
        "미국 소비 및 대외수지 상태",
        "달러 방향과 강도",
        "글로벌 자금 흐름 (미국 집중 / 분산 / 신흥국 유입)",
        "위험자산 환경 (리스크온 / 중립 / 리스크오프)",
        "한국시장에 미칠 영향",
        "한 줄 요약",
    ],
    "credit": [
        "현재 신용 스트레스 수준 (낮음 / 보통 / 높음)",
        "기업 자금조달 환경 (완화 / 중립 / 악화)",
        "유동성 환경 (확대 / 정체 / 축소)",
        "신용시장 기준 경기 선행 신호",
        "자산시장 영향 (주식 / 회사채 / 달러 / 원자재)",
        "한 줄 요약",
    ],
    "sentiment": [
        "현재 투자심리 상태 (낙관 / 중립 / 경계 / 공포)",
        "변동성 체제 변화 (안정 / 전환 / 불안정)",
        "군중심리 과열·과매도 여부",
        "소비심리 관점의 실물경제 시사점",
        "자산시장 영향 (성장주 / 가치주 / 채권 / 달러)",
        "한 줄 요약",
    ],
}

CATEGORY_GUIDELINES = {
    "business": "- PMI 50 이상/이하 기준, 전월 대비 방향 중심 판단\n- 소비 둔화 연속성 반영, GDP는 후행 보조 참고\n- 숫자 하나에 과신 금지",
    "employment": "- NFP 둔화/실업률/청구건수 방향을 함께 판단\n- 임금 MoM 반복 압력 반영\n- 항상 금리 기대와 연결",
    "inflation": "- YoY는 큰 방향, MoM은 단기 속도\n- Core 끈적함, 서비스/주거 요인 반영\n- 한 달 수치 과신 금지",
    "interest": "- 기준금리/2Y/10Y/장단기/실질금리 종합 판단\n- 경기/물가 흐름과 연결\n- 단일 금리 수치 과신 금지",
    "trade": "- 무역수지/경상수지/달러 지표를 자금 흐름과 연결\n- 달러 강세와 위험자산 환경 함께 해석\n- 단일 수치 과신 금지",
    "credit": "- 스프레드/FCI/M2로 신용 스트레스·유동성 판단\n- 경기·금리와 동시 해석\n- 단일 수치 과신 금지",
    "sentiment": "- VIX/Put-Call/심리지표를 결합 판단\n- 역지표 성격과 체제 전환 가능성 반영\n- 단일 수치 과신 금지",
}


def _to_float(value):
    """문자/숫자 값을 float로 변환. 실패 시 None"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            cleaned = value.replace("%", "").replace("K", "000").replace(",", "").strip()
            return float(cleaned)
        except Exception:
            return None
    return None


def _days_old(release_date_str):
    if not release_date_str:
        return None
    try:
        release_date = datetime.strptime(str(release_date_str), "%Y-%m-%d")
        return max((datetime.now() - release_date).days, 0)
    except Exception:
        return None


def _freshness_weight(days_old):
    if days_old is None:
        return 0.2
    if days_old <= 7:
        return 1.0
    if days_old <= 30:
        return 0.6
    if days_old <= 90:
        return 0.25
    return 0.1


def _history_trend(history_rows):
    actuals = []
    for row in history_rows:
        value = _to_float(row.get("actual"))
        if value is not None:
            actuals.append(value)

    if len(actuals) < 2:
        return {"direction": "unknown", "acceleration": "unknown"}

    recent = actuals[0]
    prev = actuals[1]
    delta = recent - prev
    direction = "up" if delta > 0 else "down" if delta < 0 else "flat"

    acceleration = "unknown"
    if len(actuals) >= 4:
        latest_move = actuals[0] - actuals[1]
        previous_move = actuals[1] - actuals[2]
        if abs(latest_move) > abs(previous_move):
            acceleration = "faster"
        elif abs(latest_move) < abs(previous_move):
            acceleration = "slower"
        else:
            acceleration = "flat"

    return {"direction": direction, "acceleration": acceleration}


def _extract_output_text(response_json):
    """Responses API 응답에서 텍스트 추출"""
    output_text = response_json.get("output_text")
    if output_text:
        return output_text

    output = response_json.get("output", [])
    for item in output:
        for content in item.get("content", []):
            if content.get("type") == "output_text" and content.get("text"):
                return content.get("text")
    return None


def _build_fallback_category_interpretation(category_summary):
    """OpenAI 실패 시 카테고리별 기본 해석 생성"""
    fallback_categories = {}
    for category, payload in category_summary.items():
        label = CATEGORY_LABELS.get(category, category)
        count = payload.get("count", 0)
        avg_surprise = payload.get("avg_surprise")
        missing = payload.get("missing_count", 0)
        freshness_score = payload.get("freshness_score", 0)
        section_titles = CATEGORY_SECTION_TITLES.get(category, ["해석", "한 줄 요약"])

        if count == 0:
            base_text = f"{label} 지표 데이터가 부족해 추세 판단이 어렵습니다."
            risk_level = "unknown"
            signals = ["데이터 업데이트 필요"]
        else:
            if avg_surprise is None:
                base_text = f"{label} 지표는 발표값이 누적되고 있으나 서프라이즈 정보가 제한적입니다."
                risk_level = "neutral"
                signals = ["발표/예측치 축적 필요"]
            elif avg_surprise > 0:
                base_text = f"{label} 지표는 평균적으로 예상치를 상회하며 상대적 개선 흐름입니다."
                risk_level = "positive"
                signals = [f"평균 서프라이즈 +{avg_surprise:.2f}"]
            elif avg_surprise < 0:
                base_text = f"{label} 지표는 평균적으로 예상치를 하회해 둔화 신호가 있습니다."
                risk_level = "caution"
                signals = [f"평균 서프라이즈 {avg_surprise:.2f}"]
            else:
                base_text = f"{label} 지표는 평균적으로 예상과 유사한 흐름입니다."
                risk_level = "neutral"
                signals = ["서프라이즈 중립"]

        if missing > 0:
            signals.append(f"미집계 지표 {missing}개")
        if freshness_score < 45:
            signals.append("신선도 낮음: 최근 데이터 부족")

        sections = []
        for idx, title in enumerate(section_titles):
            if idx == len(section_titles) - 1:
                content = f"{label}: {base_text}"
            else:
                content = base_text
            sections.append({"title": title, "content": content})

        fallback_categories[category] = {
            "label": label,
            "sections": sections,
            "signals": signals,
            "risk_level": risk_level,
            "freshness_score": freshness_score,
            "confidence": max(min(int(freshness_score), 100), 0),
            "one_line_summary": f"{label}: {base_text}",
        }

    return {
        "overall_summary": "일부 지표 기준 기본 해석입니다. OpenAI 연결 시 더 정교한 맥락 해석이 제공됩니다.",
        "categories": fallback_categories,
        "source": "fallback",
    }


def _generate_ai_indicator_interpretation(category_summary):
    """OpenAI를 사용해 카테고리별 지표 해석 생성"""
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return _build_fallback_category_interpretation(category_summary)

    category_schema = {
        category: {
            "type": "object",
            "properties": {
                "label": {"type": "string"},
                "risk_level": {
                    "type": "string",
                    "enum": ["positive", "neutral", "caution", "unknown"]
                },
                "confidence": {"type": "number"},
                "freshness_score": {"type": "number"},
                "signals": {
                    "type": "array",
                    "items": {"type": "string"}
                },
                "sections": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "title": {"type": "string"},
                            "content": {"type": "string"}
                        },
                        "required": ["title", "content"],
                        "additionalProperties": False
                    }
                },
                "one_line_summary": {"type": "string"}
            },
            "required": [
                "label",
                "risk_level",
                "confidence",
                "freshness_score",
                "signals",
                "sections",
                "one_line_summary"
            ],
            "additionalProperties": False
        } for category in CATEGORY_LABELS.keys()
    }

    schema = {
        "type": "object",
        "properties": {
            "overall_summary": {"type": "string"},
            "categories": {
                "type": "object",
                "properties": category_schema,
                "required": list(CATEGORY_LABELS.keys()),
                "additionalProperties": False
            }
        },
        "required": ["overall_summary", "categories"],
        "additionalProperties": False
    }

    frame_guide = []
    for category, titles in CATEGORY_SECTION_TITLES.items():
        frame_guide.append(
            f"[{CATEGORY_LABELS[category]}]\n"
            f"- 섹션 제목 고정: {titles}\n"
            f"- 해석 기준: {CATEGORY_GUIDELINES.get(category, '')}"
        )

    prompt = f"""당신은 매크로 전략가입니다.
아래는 DB 저장 지표를 카테고리별로 정리한 데이터입니다.
manual_check 지표는 제외되어 있습니다.

[입력 데이터]
{json.dumps(category_summary, ensure_ascii=False)}

[카테고리별 해석 프레임]
{chr(10).join(frame_guide)}

[지시사항]
1) 각 카테고리의 sections는 해당 카테고리 고정 제목 순서 그대로 작성하세요.
2) 각 section.content는 초보자도 이해 가능한 한국어로 1~3문장.
3) 숫자 하나에 과도한 확신 금지, 데이터 부족 시 판단 유보를 명시.
4) freshness_score가 낮으면 signals에 신선도 경고를 포함.
5) one_line_summary는 카테고리 핵심 결론 한 줄.
6) overall_summary는 전체 환경을 2~3문장으로 요약.
"""

    try:
        resp = requests.post(
            "https://api.openai.com/v1/responses",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": "gpt-4o-mini",
                "input": prompt,
                "text": {
                    "format": {
                        "type": "json_schema",
                        "name": "indicator_category_interpretation",
                        "strict": True,
                        "schema": schema,
                    }
                },
                "temperature": 0.4,
                "max_output_tokens": 900,
            },
            timeout=30,
        )
        if not resp.ok:
            raise RuntimeError(f"OpenAI API error: {resp.status_code} {resp.text[:300]}")

        data = resp.json()
        output_text = _extract_output_text(data)
        if not output_text:
            raise RuntimeError("OpenAI output_text is empty")

        parsed = json.loads(output_text)
        parsed["source"] = "openai"
        return parsed
    except Exception as e:
        print(f"⚠️ OpenAI interpretation failed, fallback used: {e}")
        return _build_fallback_category_interpretation(category_summary)

@app.route('/api/v2/indicators')
def get_all_indicators_from_db():
    """데이터베이스에서 모든 지표 데이터 조회 (빠른 로딩용) - 히스토리 포함"""
    try:
        force_refresh = request.args.get("force") == "1"
        now_ts = time.time()
        # 히스토리 개수 제한 (기본 12개). 0이면 히스토리 스킵.
        try:
            history_limit = int(request.args.get("history_limit", "12"))  # 기본: 12개 히스토리
        except ValueError:
            history_limit = 12

        # ✅ Phase 3: Redis 캐시 확인 (최우선)
        cache_key = f'indicators:all:v2:history_{history_limit}'
        if not force_refresh and redis_client:
            try:
                cached = redis_client.get(cache_key)
                if cached:
                    print(f"✅ Redis cache hit: {cache_key}")
                    return jsonify(json.loads(cached))
            except Exception as e:
                print(f"⚠️ Redis get error: {e}")

        # ✅ 메모리 캐시 확인 (Redis 없을 때 폴백)
        if not force_refresh and INDICATORS_CACHE["data"] and (now_ts - INDICATORS_CACHE["timestamp"] < INDICATORS_CACHE_TTL):
            return jsonify(INDICATORS_CACHE["data"])

        # indicators_config.py에서 활성화된 모든 지표 가져오기
        from crawlers.indicators_config import get_all_enabled_indicators
        all_indicator_ids = list(get_all_enabled_indicators().keys())
        results = []
        last_updated = None

        # ✅ Phase 2 최적화: 배치 쿼리 (47회 → 4회 쿼리로 감소)
        # 1. 모든 지표 데이터를 한 번에 조회
        all_data = db_service.get_multiple_indicators_data(all_indicator_ids)

        # 2. 히스토리 데이터도 한 번에 조회 (필요한 경우)
        all_history = {}
        if history_limit > 0:
            all_history = db_service.get_multiple_history_data(all_indicator_ids, limit=history_limit)

        # 3. 조회한 데이터로 results 구성
        for indicator_id in all_indicator_ids:
            data = all_data.get(indicator_id, {})
            has_error = "error" in data

            # 히스토리 데이터 가져오기
            history = all_history.get(indicator_id, [])

            # 메타데이터 추가 - 안전한 방식으로 get_indicator_config 사용
            metadata = get_indicator_config(indicator_id)

            # last_updated 추출 (가장 최신 업데이트 시간)
            indicator_updated = data.get("last_updated") if not has_error else None
            if indicator_updated:
                if not last_updated or indicator_updated > last_updated:
                    last_updated = indicator_updated

            # 해석 메타데이터 추가
            interpretation = IndicatorMetadata.get_interpretation(indicator_id)

            # Surprise 계산 (actual - forecast)
            latest = data.get("latest_release", {}) if not has_error else {}
            next_release = data.get("next_release", {}) if not has_error else {}
            actual = latest.get("actual")
            forecast = latest.get("forecast")
            surprise = None

            if actual is not None and forecast is not None:
                try:
                    # 문자열 처리 (%, K 단위)
                    actual_num = float(str(actual).replace('%', '').replace('K', '000')) if isinstance(actual, str) else float(actual)
                    forecast_num = float(str(forecast).replace('%', '').replace('K', '000')) if isinstance(forecast, str) else float(forecast)
                    surprise = round(actual_num - forecast_num, 2)
                except (ValueError, TypeError):
                    surprise = None

            results.append({
                "indicator_id": indicator_id,
                "name": CrawlerService.get_indicator_name(indicator_id),
                "name_ko": metadata.name_ko if metadata else None,
                "category": metadata.category if metadata else "business",
                "reverse_color": metadata.reverse_color if metadata else False,
                "manual_check": metadata.manual_check if metadata else False,  # 직접 확인 필요 여부
                "url": metadata.url if metadata else None,  # 직접 확인 URL
                "surprise": surprise,  # Surprise 값 추가
                "interpretation": interpretation,  # 5개 섹션 해석 추가
                "data": {
                    "latest_release": latest,
                    "next_release": next_release,
                    "history_table": history
                }
            })

        # ✅ results 배열을 딕셔너리로 변환 (중복 DB 조회 방지)
        indicators_dict = {}
        for item in results:
            indicator_id = item.get('indicator_id')
            latest_release = item.get('data', {}).get('latest_release') if item.get('data') else None
            if indicator_id and latest_release:
                # 전체 item을 저장 (latest_release뿐만 아니라 next_release, history도 포함)
                indicators_dict[indicator_id] = item

        # 3대 사이클 계산 (✅ DB 재조회 없이 results 데이터 재사용)
        try:
            macro_cycle_service = MacroCycleService(db_service)
            macro_cycle = macro_cycle_service.calculate_cycle_from_data(indicators_dict)
        except Exception as e:
            print(f"Macro cycle calculation error: {e}")
            macro_cycle = None

        try:
            credit_cycle_service = CreditCycleService(db_service)
            credit_cycle = credit_cycle_service.calculate_cycle_from_data(indicators_dict)
        except Exception as e:
            print(f"Credit cycle calculation error: {e}")
            credit_cycle = None

        try:
            from services.sentiment_cycle_service import SentimentCycleService
            sentiment_cycle_service = SentimentCycleService(db_service)
            sentiment_cycle = sentiment_cycle_service.calculate_cycle_from_data(indicators_dict)
        except Exception as e:
            print(f"Sentiment cycle calculation error: {e}")
            sentiment_cycle = None

        # ✅ Master Market Cycle 계산 (3대 사이클 통합)
        # Phase 1 최적화: 이미 조회한 indicators_dict 재사용 (DB 재조회 17회 제거)
        master_cycle = None
        try:
            from services.cycle_engine import calculate_master_cycle_v1_from_data
            master_cycle = calculate_master_cycle_v1_from_data(indicators_dict)
        except Exception as e:
            print(f"Master cycle calculation error: {e}")
            master_cycle = None

        response_data = {
            "status": "success",
            "indicators": results,
            "total_count": len(results),
            "source": "database",
            "last_updated": last_updated,  # 가장 최신 업데이트 시간
            # ✅ 3대 사이클 데이터 추가 (4개 요청 → 1개 요청 최적화)
            "macro_cycle": macro_cycle,
            "credit_cycle": credit_cycle,
            "sentiment_cycle": sentiment_cycle,
            # ✅ Master Market Cycle 추가 (3대 사이클 통합 점수)
            "master_cycle": master_cycle
        }

        # ✅ 캐시 저장
        INDICATORS_CACHE["data"] = response_data
        INDICATORS_CACHE["timestamp"] = now_ts

        # ✅ Phase 3: Redis 캐시 저장 (5분 TTL)
        if redis_client:
            try:
                redis_client.setex(
                    cache_key,
                    300,  # 5분 TTL
                    json.dumps(response_data)
                )
                print(f"✅ Redis cache saved: {cache_key}")
            except Exception as e:
                print(f"⚠️ Redis set error: {e}")

        return jsonify(response_data)

    except Exception as e:
        import traceback
        print(f"Error in get_all_indicators_from_db: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"Database query failed: {str(e)}"
        }), 500


@app.route('/api/v2/indicators/ai-interpretation', methods=['GET'])
def get_ai_interpretation_for_indicators():
    """저장된 모든 지표 수치 기반 카테고리별 AI 종합 해석"""
    try:
        all_indicator_ids = list(get_all_enabled_indicators().keys())
        if hasattr(db_service, "get_multiple_indicators_data"):
            all_data = db_service.get_multiple_indicators_data(all_indicator_ids)
        else:
            all_data = {indicator_id: db_service.get_indicator_data(indicator_id) for indicator_id in all_indicator_ids}

        if hasattr(db_service, "get_multiple_history_data"):
            all_history = db_service.get_multiple_history_data(all_indicator_ids, limit=6)
        else:
            all_history = {
                indicator_id: db_service.get_history_data(indicator_id, limit=6)
                for indicator_id in all_indicator_ids
            }

        category_details = {key: [] for key in CATEGORY_LABELS.keys()}
        excluded_manual_check = []

        for indicator_id in all_indicator_ids:
            metadata = get_indicator_config(indicator_id)
            if not metadata:
                continue
            if metadata.manual_check:
                excluded_manual_check.append(indicator_id)
                continue

            category = metadata.category
            if category not in category_details:
                continue

            data = all_data.get(indicator_id, {})
            latest = data.get("latest_release", {}) if isinstance(data, dict) else {}
            history_rows = all_history.get(indicator_id, []) if isinstance(all_history, dict) else []

            actual = latest.get("actual")
            forecast = latest.get("forecast")
            previous = latest.get("previous")
            release_date = latest.get("release_date")
            days_old = _days_old(release_date)
            freshness_weight = _freshness_weight(days_old)
            trend = _history_trend(history_rows)

            actual_num = _to_float(actual)
            forecast_num = _to_float(forecast)
            previous_num = _to_float(previous)

            surprise = None
            if actual_num is not None and forecast_num is not None:
                surprise = round(actual_num - forecast_num, 2)

            delta_prev = None
            if actual_num is not None and previous_num is not None:
                delta_prev = round(actual_num - previous_num, 2)

            category_details[category].append({
                "indicator_id": indicator_id,
                "name_ko": metadata.name_ko,
                "name": metadata.name,
                "actual": actual,
                "forecast": forecast,
                "previous": previous,
                "surprise": surprise,
                "delta_prev": delta_prev,
                "release_date": release_date,
                "days_old": days_old,
                "freshness_weight": freshness_weight,
                "trend_direction": trend.get("direction"),
                "trend_acceleration": trend.get("acceleration"),
                "history_actuals": [
                    _to_float(row.get("actual")) for row in history_rows[:6]
                ],
            })

        category_summary = {}
        for category, rows in category_details.items():
            valid_surprises = [row["surprise"] for row in rows if row["surprise"] is not None]
            weighted_sum = sum((row.get("freshness_weight", 0.2) * 100) for row in rows)
            freshness_score = round(weighted_sum / len(rows), 1) if rows else 0

            improving_count = len([row for row in rows if row.get("trend_direction") == "up"])
            weakening_count = len([row for row in rows if row.get("trend_direction") == "down"])
            category_summary[category] = {
                "label": CATEGORY_LABELS[category],
                "count": len(rows),
                "avg_surprise": round(sum(valid_surprises) / len(valid_surprises), 2) if valid_surprises else None,
                "missing_count": sum(1 for row in rows if row["actual"] in [None, "-", ""]),
                "freshness_score": freshness_score,
                "improving_count": improving_count,
                "weakening_count": weakening_count,
                "indicators": rows[:12],  # 프롬프트 길이 제어
            }

        interpretation = _generate_ai_indicator_interpretation(category_summary)

        return jsonify({
            "status": "success",
            "generated_at": datetime.now().isoformat(),
            "source": interpretation.get("source", "fallback"),
            "overall_summary": interpretation.get("overall_summary"),
            "categories": interpretation.get("categories"),
            "category_summary": category_summary,
            "excluded_manual_check_count": len(excluded_manual_check),
            "excluded_manual_check_ids": excluded_manual_check,
        })
    except Exception as e:
        import traceback
        print(f"Error in ai interpretation endpoint: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"AI interpretation failed: {str(e)}"
        }), 500

@app.route('/api/v2/indicators/health-check')
def get_indicators_health_check():
    """모든 지표의 데이터 신선도 및 상태 확인"""
    try:
        from datetime import datetime, timedelta
        from crawlers.indicators_config import get_all_enabled_indicators, get_indicator_config

        all_indicator_ids = list(get_all_enabled_indicators().keys())
        health_results = []
        now = datetime.now()

        # 상태별 카운터
        counts = {
            "healthy": 0,
            "stale": 0,
            "outdated": 0,
            "manual_check": 0,
            "updated_recent": 0,
            "error": 0
        }

        for indicator_id in all_indicator_ids:
            metadata = get_indicator_config(indicator_id)
            if metadata and metadata.manual_check:
                health_results.append({
                    "indicator_id": indicator_id,
                    "name": CrawlerService.get_indicator_name(indicator_id),
                    "status": "manual_check",
                    "last_update": None,
                    "days_old": None,
                    "message": "직접 확인 필요"
                })
                counts["manual_check"] += 1
                continue

            # 지표 데이터 조회
            data = db_service.get_indicator_data(indicator_id)

            crawl_info = db_service.get_crawl_info(indicator_id)
            crawl_error = crawl_info and crawl_info.get("status") == "error"
            error_message = crawl_info.get("error_message") if crawl_info else None

            if "error" in data:
                # 데이터 조회 오류
                health_results.append({
                    "indicator_id": indicator_id,
                    "name": CrawlerService.get_indicator_name(indicator_id),
                    "status": "error",
                    "last_update": None,
                    "days_old": None,
                    "message": error_message or "데이터 조회 실패"
                })
                counts["error"] += 1
                continue

            latest = data.get("latest_release", {})
            release_date_str = latest.get("release_date")

            if not release_date_str or release_date_str == "미정":
                # 날짜 정보 없음
                health_results.append({
                    "indicator_id": indicator_id,
                    "name": CrawlerService.get_indicator_name(indicator_id),
                    "status": "error",
                    "last_update": release_date_str or "없음",
                    "days_old": None,
                    "message": "날짜 정보 없음"
                })
                counts["error"] += 1
                continue

            try:
                # 날짜/시간 파싱 (시간 없으면 00:00)
                release_time_str = latest.get("time")
                release_dt = datetime.strptime(release_date_str, "%Y-%m-%d")
                if release_time_str and release_time_str != "미정":
                    for fmt in ("%H:%M", "%H:%M:%S"):
                        try:
                            t = datetime.strptime(release_time_str, fmt).time()
                            release_dt = release_dt.replace(hour=t.hour, minute=t.minute, second=t.second)
                            break
                        except ValueError:
                            continue
                release_date = release_dt

                last_crawl_time = crawl_info.get("last_crawl_time") if crawl_info else None
                last_crawl_dt = None
                if last_crawl_time:
                    try:
                        if isinstance(last_crawl_time, str):
                            try:
                                parsed = datetime.strptime(last_crawl_time, "%Y-%m-%d %H:%M:%S")
                            except ValueError:
                                parsed = datetime.strptime(last_crawl_time, "%a, %d %b %Y %H:%M:%S GMT")
                            last_crawl_dt = parsed
                        else:
                            last_crawl_dt = last_crawl_time
                    except Exception:
                        last_crawl_dt = None

                # 크롤링 시간이 없으면 release_date 기준으로 경과일 계산
                if last_crawl_dt:
                    delta = abs(last_crawl_dt - release_date)
                else:
                    delta = abs(now - release_date)
                days_old = delta.days

                if delta <= timedelta(hours=24):
                    counts["updated_recent"] += 1
                    status = "healthy"
                    message = "최근 데이터"
                    counts["healthy"] += 1
                elif days_old <= 7:
                    status = "healthy"
                    message = "최신 데이터"
                    counts["healthy"] += 1
                elif days_old <= 30:
                    status = "stale"
                    message = "약간 오래된 데이터"
                    counts["stale"] += 1
                else:
                    status = "outdated"
                    message = "매우 오래된 데이터"
                    counts["outdated"] += 1

                if crawl_error:
                    message = f"{message} (크롤링 실패 기록 있음)"

                health_results.append({
                    "indicator_id": indicator_id,
                    "name": CrawlerService.get_indicator_name(indicator_id),
                    "status": status,
                    "last_update": release_date_str,
                    "days_old": days_old,
                    "message": message
                })

            except ValueError:
                # 날짜 파싱 실패
                health_results.append({
                    "indicator_id": indicator_id,
                    "name": CrawlerService.get_indicator_name(indicator_id),
                    "status": "error",
                    "last_update": release_date_str,
                    "days_old": None,
                    "message": "날짜 형식 오류"
                })
                counts["error"] += 1

        # 상태별 정렬 (error > outdated > stale > healthy)
        status_priority = {"error": 0, "manual_check": 1, "outdated": 2, "stale": 3, "healthy": 4}
        health_results.sort(key=lambda x: (status_priority.get(x["status"], 4), x["days_old"] if x["days_old"] is not None else 999))

        return jsonify({
            "status": "success",
            "timestamp": now.isoformat(),
            "total_indicators": len(all_indicator_ids),
            "summary": counts,
            "indicators": health_results
        })

    except Exception as e:
        import traceback
        print(f"Error in get_indicators_health_check: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"Health check failed: {str(e)}"
        }), 500

@app.route('/api/v2/indicators/<indicator_id>')
def get_indicator_from_db(indicator_id):
    """데이터베이스에서 특정 지표 데이터 조회"""
    try:
        data = db_service.get_indicator_data(indicator_id)

        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 404

        return jsonify({
            "status": "success",
            "indicator": CrawlerService.get_indicator_name(indicator_id),
            "data": data,
            "source": "database"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Database query failed: {str(e)}"
        }), 500


# ========================================
# 경제지표 사용자 해석 API
# ========================================

@app.route('/api/v2/indicator-interpretation/<indicator_id>', methods=['GET'])
def get_indicator_interpretation(indicator_id):
    """특정 지표의 사용자 해석 조회"""
    try:
        interpretation = db_service.get_indicator_interpretation(indicator_id)
        return jsonify({
            "status": "success",
            "indicator_id": indicator_id,
            "user_interpretation": interpretation
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Failed to get interpretation: {str(e)}"
        }), 500


@app.route('/api/v2/indicator-interpretation/<indicator_id>', methods=['PUT'])
def save_indicator_interpretation(indicator_id):
    """경제지표 사용자 해석 저장"""
    try:
        data = request.get_json()
        if not data or 'user_interpretation' not in data:
            return jsonify({
                "status": "error",
                "message": "user_interpretation field is required"
            }), 400

        user_interpretation = data['user_interpretation']
        success = db_service.save_indicator_interpretation(indicator_id, user_interpretation)

        if success:
            return jsonify({
                "status": "success",
                "indicator_id": indicator_id,
                "message": "Interpretation saved successfully"
            })
        else:
            return jsonify({
                "status": "error",
                "message": "Failed to save interpretation"
            }), 500

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Failed to save interpretation: {str(e)}"
        }), 500


@app.route('/api/v2/history/<indicator_id>')
def get_history_from_db(indicator_id):
    """데이터베이스에서 히스토리 데이터 조회"""
    try:
        try:
            limit = int(request.args.get("limit")) if request.args.get("limit") else None
        except ValueError:
            limit = None

        history_data = db_service.get_history_data(indicator_id, limit=limit)

        return jsonify({
            "status": "success",
            "indicator": indicator_id,
            "data": history_data,
            "total_rows": len(history_data),
            "source": "database"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Database query failed: {str(e)}"
        }), 500

async def update_all_indicators_background_async():
    """백그라운드에서 모든 지표 업데이트 실행 (비동기 병렬 크롤링)"""
    global update_status, INDICATORS_CACHE
    import asyncio

    try:
        update_status["is_updating"] = True
        update_status["start_time"] = time.time()
        update_status["progress"] = 0
        update_status["completed_indicators"] = []
        update_status["failed_indicators"] = []
        update_status["current_indicator"] = ""
        update_status["total_indicators"] = 0
        save_update_status()

        # indicators_config.py에서 활성화된 모든 지표 사용 (manual_check 제외)
        from crawlers.indicators_config import get_all_enabled_indicators
        indicators = [
            indicator_id
            for indicator_id, config in get_all_enabled_indicators().items()
            if not config.manual_check
        ]
        total_indicators = len(indicators)
        update_status["total_indicators"] = total_indicators  # 전체 개수 저장
        save_update_status()

        # 비동기 크롤링 설정
        # 동시 요청 제한 + 지터로 429 완화
        max_concurrency = int(os.getenv("CRAWL_MAX_CONCURRENCY", "4"))
        semaphore = asyncio.Semaphore(max_concurrency)

        async def run_indicator(indicator_id: str):
            async with semaphore:
                await asyncio.sleep(0.2 + (hash(indicator_id) % 5) * 0.1)
                result = await asyncio.to_thread(CrawlerService.crawl_indicator, indicator_id)
                return indicator_id, result

        tasks = [asyncio.create_task(run_indicator(indicator_id)) for indicator_id in indicators]

        # 완료되는 순서대로 결과 처리 (진행률 실시간 반영)
        completed_count = 0
        for task in asyncio.as_completed(tasks):
            indicator_id, result = await task

            if isinstance(result, Exception):
                db_service.update_crawl_info(
                    indicator_id,
                    "error",
                    0,
                    f"Exception: {str(result)}"
                )
                update_status["failed_indicators"].append({
                    "indicator_id": indicator_id,
                    "error": f"Exception: {str(result)}"
                })
            elif isinstance(result, dict) and "error" in result:
                db_service.update_crawl_info(
                    indicator_id,
                    "error",
                    0,
                    result["error"]
                )
                update_status["failed_indicators"].append({
                    "indicator_id": indicator_id,
                    "error": result["error"]
                })
            else:
                # 데이터베이스에 저장
                db_service.save_indicator_data(indicator_id, result)
                update_status["completed_indicators"].append(indicator_id)

            completed_count += 1
            update_status["progress"] = int((completed_count / total_indicators) * 100)
            update_status["current_indicator"] = indicator_id
            save_update_status()

        update_status["progress"] = 100
        update_status["current_indicator"] = ""
        save_update_status()

    except Exception as e:
        import traceback
        update_status["failed_indicators"].append({
            "indicator_id": "system",
            "error": f"Update process failed: {str(e)}\n{traceback.format_exc()}"
        })
        save_update_status()
    finally:
        # ✅ 지표 업데이트 후 캐시 무효화 + 상태 리셋
        INDICATORS_CACHE["data"] = None
        INDICATORS_CACHE["timestamp"] = 0

        # ✅ Phase 3: Redis 캐시 무효화 (모든 history_limit 키 삭제)
        if redis_client:
            try:
                # 패턴 매칭으로 모든 indicators 캐시 키 삭제
                for key in redis_client.scan_iter("indicators:all:v2:*"):
                    redis_client.delete(key)
                print("✅ Redis cache invalidated")
            except Exception as e:
                print(f"⚠️ Redis cache invalidation error: {e}")

        update_status["is_updating"] = False
        save_update_status()

def update_all_indicators_background():
    """비동기 함수를 동기 컨텍스트에서 실행하는 래퍼"""
    import asyncio

    # 새 이벤트 루프 생성 및 실행
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(update_all_indicators_background_async())
    finally:
        loop.close()

@app.route('/api/v2/update-indicators', methods=['POST'])
def trigger_update_indicators():
    """모든 지표 업데이트 트리거 (백그라운드 실행)"""
    global update_status

    current_status = load_update_status()
    if current_status:
        update_status = current_status

    # 이전 업데이트가 비정상적으로 오래 걸리는 경우 스테일 처리 (10분 초과 시 리셋)
    if update_status["is_updating"]:
        now_ts = time.time()
        start_ts = update_status.get("start_time") or 0
        if start_ts and (now_ts - start_ts) > MAX_UPDATE_DURATION:
            update_status["is_updating"] = False
            update_status["progress"] = 0
            update_status["current_indicator"] = ""
            update_status["completed_indicators"] = []
            update_status["failed_indicators"] = []
            save_update_status()
        else:
            return jsonify({
                "status": "error",
                "message": "Update is already in progress"
            }), 409

    # 백그라운드 스레드로 업데이트 실행
    thread = threading.Thread(target=update_all_indicators_background)
    thread.daemon = True
    thread.start()

    # 업데이트 시작 상태를 즉시 반영 (상태 폴링 레이스 방지)
    update_status["is_updating"] = True
    update_status["start_time"] = time.time()
    update_status["progress"] = 0
    update_status["current_indicator"] = ""
    save_update_status()

    return jsonify({
        "status": "success",
        "message": "Update started in background",
        "check_status_url": "/api/v2/update-status"
    })

@app.route('/api/v2/update-status')
def get_update_status():
    """업데이트 진행 상황 조회"""
    current_status = load_update_status()
    return jsonify({
        "status": "success",
        "update_status": current_status
    })

@app.route('/api/v2/reset-update-status', methods=['POST'])
def reset_update_status():
    """업데이트 상태 강제 리셋 (stuck 상태 해결용)"""
    global update_status
    update_status["is_updating"] = False
    update_status["current_indicator"] = ""
    update_status["progress"] = 100
    update_status["completed_indicators"] = []
    update_status["failed_indicators"] = []
    update_status["start_time"] = None
    update_status["total_indicators"] = 0
    save_update_status()

    return jsonify({
        "status": "success",
        "message": "Update status has been reset",
        "update_status": update_status
    })

@app.route('/api/v2/crawl-info')
def get_all_crawl_info():
    """모든 지표의 크롤링 정보 조회"""
    try:
        # enabled 지표 기준으로 crawl_info 수집 (indicators 테이블 미시드 대비)
        indicators = list(get_all_enabled_indicators().keys())
        crawl_info_list = []

        for indicator_id in indicators:
            crawl_info = db_service.get_crawl_info(indicator_id)
            if crawl_info:
                crawl_info_list.append(crawl_info)
            else:
                crawl_info_list.append({
                    "indicator_id": indicator_id,
                    "last_crawl_time": None,
                    "status": "missing",
                    "error_message": None,
                    "data_count": 0
                })

        return jsonify({
            "status": "success",
            "crawl_info": crawl_info_list
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

@app.route('/api/v2/crawl/<indicator_id>', methods=['POST'])
def crawl_single_indicator(indicator_id):
    """단일 지표 크롤링 후 데이터베이스 저장"""
    try:
        # 크롤링 실행
        crawled_data = CrawlerService.crawl_indicator(indicator_id)

        if "error" in crawled_data:
            return jsonify({
                "status": "error",
                "message": crawled_data["error"]
            }), 400

        # 데이터베이스에 저장
        success = db_service.save_indicator_data(indicator_id, crawled_data)

        if success:
            return jsonify({
                "status": "success",
                "message": f"Successfully crawled and saved {indicator_id}",
                "data": crawled_data
            })
        else:
            return jsonify({
                "status": "error",
                "message": f"Failed to save {indicator_id} to database"
            }), 500

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Error crawling {indicator_id}: {str(e)}"
        }), 500

@app.route('/api/add-asset', methods=['POST'])
def add_asset():
    """포트폴리오 자산 추가 API"""
    try:
        # JSON 데이터 받기
        data = request.get_json()
        print(f"Received data: {data}")

        if not data:
            return jsonify({
                "status": "error",
                "message": "No data provided"
            }), 400

        # 필수 필드 검증
        required_fields = ['assetType', 'name', 'date', 'user_id']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({
                    "status": "error",
                    "message": f"Missing required field: {field}"
                }), 400

        # 데이터 변환 (프론트엔드 형식 → DB 형식)
        asset_data = {
            'user_id': data.get('user_id'),
            'asset_type': data.get('assetType'),
            'sub_category': data.get('subCategory'),
            'name': data.get('name'),
            'amount': data.get('amount'),
            'quantity': data.get('quantity'),
            'avg_price': data.get('avgPrice'),
            'eval_amount': data.get('evaluationAmount'),
            'date': data.get('date'),
            'note': data.get('note'),

            # 소분류별 전용 필드들
            # 부동산 필드
            'area_pyeong': data.get('area_pyeong'),
            'acquisition_tax': data.get('acquisition_tax'),
            'lawyer_fee': data.get('lawyer_fee'),
            'brokerage_fee': data.get('brokerage_fee'),
            'rent_type': data.get('rent_type'),
            'rental_income': data.get('rental_income'),
            'jeonse_deposit': data.get('jeonse_deposit'),
            # 예금/적금 필드
            'maturity_date': data.get('maturity_date'),
            'interest_rate': data.get('interest_rate'),
            'early_withdrawal_fee': data.get('early_withdrawal_fee'),
            # MMF/CMA 필드
            'current_yield': data.get('current_yield'),
            'annual_yield': data.get('annual_yield'),
            'minimum_balance': data.get('minimum_balance'),
            'withdrawal_fee': data.get('withdrawal_fee'),
            # 주식/ETF 필드
            'dividend_rate': data.get('dividend_rate'),
            # 펀드 필드
            'nav': data.get('nav'),
            'management_fee': data.get('management_fee')
        }

        # 데이터베이스에 저장
        print(f"About to save asset_data: {asset_data}")
        print(f"db_service type: {type(db_service)}")
        print(f"db_service has save_asset: {hasattr(db_service, 'save_asset')}")

        result = db_service.save_asset(asset_data)
        print(f"Save result: {result}")

        if result.get('status') == 'success':
            # 포트폴리오 이력 추가 - 자산 추가
            try:
                # 현재 포트폴리오 총액 계산
                portfolio_data = db_service.get_all_assets(data.get('user_id'))
                if portfolio_data.get('status') == 'success':
                    assets = portfolio_data.get('data', [])
                    total_assets = sum(asset.get('amount', 0) for asset in assets)
                    total_principal = sum(asset.get('principal', asset.get('amount', 0)) for asset in assets)
                    total_eval_amount = sum(asset.get('eval_amount', asset.get('amount', 0)) for asset in assets)

                    # 히스토리 저장
                    change_amount = data.get('amount', 0)
                    db_service.save_portfolio_history(
                        user_id=data.get('user_id'),
                        change_type='add',
                        total_assets=total_assets,
                        total_principal=total_principal,
                        total_eval_amount=total_eval_amount,
                        asset_name=data.get('name'),
                        change_amount=change_amount,
                        notes=f"자산 추가: {data.get('name')}"
                    )
            except Exception as history_error:
                print(f"Failed to save portfolio history: {history_error}")
                # 히스토리 저장 실패해도 메인 작업은 성공으로 처리

            return jsonify(result)
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error in add_asset: {e}")
        return jsonify({
            "status": "error",
            "message": f"Server error: {str(e)}"
        }), 500

@app.route('/api/test-db', methods=['GET'])
def test_db():
    """데이터베이스 연결 및 메서드 테스트"""
    try:
        print(f"Testing database connection...")
        print(f"Database service type: {type(db_service)}")
        print(f"Database service methods: {[m for m in dir(db_service) if not m.startswith('_')]}")

        return jsonify({
            "status": "success",
            "db_type": str(type(db_service)),
            "has_save_asset": hasattr(db_service, 'save_asset')
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/portfolio', methods=['GET'])
def get_portfolio():
    """사용자별 포트폴리오 데이터 조회 API"""
    try:
        # 쿼리 파라미터에서 user_id 가져오기
        user_id = request.args.get('user_id')

        if user_id:
            user_id = int(user_id)

        # PostgreSQL에서 사용자별 자산 데이터 조회
        result = db_service.get_all_assets(user_id)

        if result.get('status') == 'success':
            return jsonify(result)
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error getting portfolio: {e}")
        return jsonify({
            "status": "error",
            "message": f"포트폴리오 조회 실패: {str(e)}"
        }), 500

@app.route('/api/portfolio/export/excel', methods=['GET'])
def export_portfolio_excel():
    """포트폴리오 데이터를 Excel 파일로 다운로드"""
    try:
        # 쿼리 파라미터에서 user_id 가져오기
        user_id = request.args.get('user_id')
        if not user_id:
            return jsonify({"status": "error", "message": "user_id required"}), 400

        user_id = int(user_id)

        # PostgreSQL에서 사용자별 자산 데이터 조회
        result = db_service.get_all_assets(user_id)

        if result.get('status') != 'success':
            return jsonify({"status": "error", "message": "데이터 조회 실패"}), 500

        assets = result.get('data', [])

        # Excel 파일 생성
        wb = Workbook()

        # Sheet 1: 자산 목록
        ws1 = wb.active
        ws1.title = "자산 목록"

        # 헤더 스타일
        header_fill = PatternFill(start_color="DAA520", end_color="DAA520", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # 헤더
        headers = ['대분류', '소분류', '자산명', '수량', '매수평균가', '투자원금', '평가금액', '수익률(%)', '등록일']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 데이터
        for row_idx, asset in enumerate(assets, 2):
            ws1.cell(row=row_idx, column=1, value=asset.get('asset_type', ''))
            ws1.cell(row=row_idx, column=2, value=asset.get('sub_category', ''))
            ws1.cell(row=row_idx, column=3, value=asset.get('name', ''))
            ws1.cell(row=row_idx, column=4, value=asset.get('quantity', 0))
            ws1.cell(row=row_idx, column=5, value=asset.get('avg_price', 0))
            ws1.cell(row=row_idx, column=6, value=asset.get('principal', asset.get('amount', 0)))
            ws1.cell(row=row_idx, column=7, value=asset.get('eval_amount', asset.get('amount', 0)))
            ws1.cell(row=row_idx, column=8, value=round(asset.get('profit_rate', 0), 2))
            ws1.cell(row=row_idx, column=9, value=asset.get('date', ''))

        # 컬럼 너비 조정
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 20
        ws1.column_dimensions['D'].width = 10
        ws1.column_dimensions['E'].width = 12
        ws1.column_dimensions['F'].width = 12
        ws1.column_dimensions['G'].width = 12
        ws1.column_dimensions['H'].width = 12
        ws1.column_dimensions['I'].width = 12

        # Sheet 2: 자산군별 요약
        ws2 = wb.create_sheet("자산군별 요약")

        # 헤더
        headers2 = ['자산군', '총액', '비중(%)', '평균수익률(%)', '자산개수']
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 자산군별 집계
        asset_summary = {}
        total_eval = sum(asset.get('eval_amount', asset.get('amount', 0)) for asset in assets)

        for asset in assets:
            asset_type = asset.get('asset_type', '기타')
            eval_amount = asset.get('eval_amount', asset.get('amount', 0))
            profit_rate = asset.get('profit_rate', 0)

            if asset_type not in asset_summary:
                asset_summary[asset_type] = {
                    'total': 0,
                    'count': 0,
                    'profit_rates': []
                }

            asset_summary[asset_type]['total'] += eval_amount
            asset_summary[asset_type]['count'] += 1
            asset_summary[asset_type]['profit_rates'].append(profit_rate)

        # 데이터
        for row_idx, (asset_type, summary) in enumerate(asset_summary.items(), 2):
            ws2.cell(row=row_idx, column=1, value=asset_type)
            ws2.cell(row=row_idx, column=2, value=summary['total'])
            ws2.cell(row=row_idx, column=3, value=round((summary['total'] / total_eval * 100) if total_eval > 0 else 0, 2))
            avg_profit = sum(summary['profit_rates']) / len(summary['profit_rates']) if summary['profit_rates'] else 0
            ws2.cell(row=row_idx, column=4, value=round(avg_profit, 2))
            ws2.cell(row=row_idx, column=5, value=summary['count'])

        # 컬럼 너비 조정
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws2.column_dimensions[col].width = 15

        # Sheet 3: 목표 달성 현황
        ws3 = wb.create_sheet("목표 달성 현황")

        # 헤더
        headers3 = ['항목', '값']
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 전체 자산 통계
        total_principal = sum(asset.get('principal', asset.get('amount', 0)) for asset in assets)
        total_eval_amount = sum(asset.get('eval_amount', asset.get('amount', 0)) for asset in assets)
        total_profit = total_eval_amount - total_principal
        total_profit_rate = ((total_eval_amount - total_principal) / total_principal * 100) if total_principal > 0 else 0

        ws3.cell(row=2, column=1, value="투자원금")
        ws3.cell(row=2, column=2, value=total_principal)
        ws3.cell(row=3, column=1, value="현재가치")
        ws3.cell(row=3, column=2, value=total_eval_amount)
        ws3.cell(row=4, column=1, value="총 수익")
        ws3.cell(row=4, column=2, value=total_profit)
        ws3.cell(row=5, column=1, value="수익률(%)")
        ws3.cell(row=5, column=2, value=round(total_profit_rate, 2))
        ws3.cell(row=6, column=1, value="자산 개수")
        ws3.cell(row=6, column=2, value=len(assets))

        # 컬럼 너비 조정
        ws3.column_dimensions['A'].width = 15
        ws3.column_dimensions['B'].width = 20

        # 파일명 생성
        filename = f"portfolio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # 메모리 버퍼에 저장
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error exporting portfolio to Excel: {e}")
        return jsonify({
            "status": "error",
            "message": f"Excel 생성 실패: {str(e)}"
        }), 500

@app.route('/api/update-asset/<int:asset_id>', methods=['PUT'])
def update_asset(asset_id):
    """포트폴리오 자산 수정 API"""
    try:
        data = request.get_json()
        print(f"Attempting to update asset with ID: {asset_id}")
        print(f"Update data: {data}")

        if data:
            if 'eval_amount' not in data:
                if 'evaluation_amount' in data:
                    data['eval_amount'] = data.get('evaluation_amount')
                elif 'evaluationAmount' in data:
                    data['eval_amount'] = data.get('evaluationAmount')

        # 필수 필드 검증
        required_fields = ['asset_type', 'name', 'date']
        for field in required_fields:
            if field not in data:
                return jsonify({
                    "status": "error",
                    "message": f"필수 필드가 누락되었습니다: {field}"
                }), 400

        # 데이터베이스에서 자산 업데이트
        result = db_service.update_asset(asset_id, data)

        if result.get('status') == 'success':
            # 포트폴리오 이력 추가 - 자산 수정
            try:
                user_id = data.get('user_id')  # user_id는 update data에 포함되어야 함
                if user_id:
                    # 현재 포트폴리오 총액 계산
                    portfolio_data = db_service.get_all_assets(user_id)
                    if portfolio_data.get('status') == 'success':
                        assets = portfolio_data.get('data', [])
                        total_assets = sum(asset.get('amount', 0) for asset in assets)
                        total_principal = sum(asset.get('principal', asset.get('amount', 0)) for asset in assets)
                        total_eval_amount = sum(asset.get('eval_amount', asset.get('amount', 0)) for asset in assets)

                        # 히스토리 저장
                        db_service.save_portfolio_history(
                            user_id=user_id,
                            change_type='update',
                            total_assets=total_assets,
                            total_principal=total_principal,
                            total_eval_amount=total_eval_amount,
                            asset_id=asset_id,
                            asset_name=data.get('name'),
                            notes=f"자산 수정: {data.get('name')}"
                        )
            except Exception as history_error:
                print(f"Failed to save portfolio history: {history_error}")
                # 히스토리 저장 실패해도 메인 작업은 성공으로 처리

            return jsonify(result)
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error updating asset: {e}")
        return jsonify({
            "status": "error",
            "message": f"자산 수정 실패: {str(e)}"
        }), 500

@app.route('/api/delete-asset/<int:asset_id>', methods=['DELETE'])
def delete_asset(asset_id):
    """포트폴리오 자산 삭제 API"""
    try:
        print(f"Attempting to delete asset with ID: {asset_id}")

        # 쿼리 파라미터에서 user_id 가져오기
        user_id = request.args.get('user_id')
        if not user_id:
            return jsonify({
                "status": "error",
                "message": "user_id is required"
            }), 400

        # 삭제 전 자산 정보 조회 (히스토리용)
        asset_info = None
        try:
            assets_data = db_service.get_all_assets(user_id)
            if assets_data.get('status') == 'success':
                assets = assets_data.get('data', [])
                asset_info = next((asset for asset in assets if asset.get('id') == asset_id), None)
        except Exception as e:
            print(f"Failed to get asset info before deletion: {e}")

        # 데이터베이스에서 자산 삭제 (user_id 검증 포함)
        result = db_service.delete_asset(asset_id, user_id)

        if result.get('status') == 'success':
            # 포트폴리오 이력 추가 - 자산 삭제
            try:
                # 현재 포트폴리오 총액 계산 (삭제 후)
                portfolio_data = db_service.get_all_assets(user_id)
                if portfolio_data.get('status') == 'success':
                    assets = portfolio_data.get('data', [])
                    total_assets = sum(asset.get('amount', 0) for asset in assets)
                    total_principal = sum(asset.get('principal', asset.get('amount', 0)) for asset in assets)
                    total_eval_amount = sum(asset.get('eval_amount', asset.get('amount', 0)) for asset in assets)

                    # 히스토리 저장
                    asset_name = asset_info.get('name', f'ID {asset_id}') if asset_info else f'ID {asset_id}'
                    change_amount = -(asset_info.get('amount', 0)) if asset_info else 0  # 음수로 표시
                    db_service.save_portfolio_history(
                        user_id=user_id,
                        change_type='delete',
                        total_assets=total_assets,
                        total_principal=total_principal,
                        total_eval_amount=total_eval_amount,
                        asset_id=asset_id,
                        asset_name=asset_name,
                        change_amount=change_amount,
                        notes=f"자산 삭제: {asset_name}"
                    )
            except Exception as history_error:
                print(f"Failed to save portfolio history: {history_error}")
                # 히스토리 저장 실패해도 메인 작업은 성공으로 처리

            return jsonify(result)
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error deleting asset: {e}")
        return jsonify({
            "status": "error",
            "message": f"자산 삭제 실패: {str(e)}"
        }), 500

@app.route('/api/goal-settings', methods=['GET'])
def get_goal_settings():
    """목표 설정 조회 API"""
    try:
        user_id = request.args.get('user_id', 'default')
        result = db_service.get_goal_settings(user_id)

        if result.get('status') == 'success':
            return jsonify(result)
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error getting goal settings: {e}")
        return jsonify({
            "status": "error",
            "message": f"목표 설정 조회 실패: {str(e)}"
        }), 500

@app.route('/api/goal-settings', methods=['POST'])
def save_goal_settings():
    """목표 설정 저장 API"""
    try:
        data = request.get_json()
        user_id = data.get('user_id', 'default')
        total_goal = data.get('total_goal', 50000000)
        target_date = data.get('target_date', '2024-12-31')
        category_goals = data.get('category_goals', {})
        sub_category_goals = data.get('sub_category_goals', {})

        print(f"Saving goal settings: user_id={user_id}, total_goal={total_goal}, target_date={target_date}")

        result = db_service.save_goal_settings(user_id, total_goal, target_date, category_goals, sub_category_goals)

        if result.get('status') == 'success':
            return jsonify(result)
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error saving goal settings: {e}")
        return jsonify({
            "status": "error",
            "message": f"목표 설정 저장 실패: {str(e)}"
        }), 500

# === 가계부 예산 목표 설정 API ===

@app.route('/api/expense-budget-goals', methods=['GET'])
def get_expense_budget_goals():
    """가계부 예산 목표 조회 API"""
    try:
        user_id = request.args.get('user_id', 'default')
        result = db_service.get_expense_budget_goals(user_id)

        if result.get('status') == 'success':
            return jsonify(result)
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error getting expense budget goals: {e}")
        return jsonify({
            "status": "error",
            "message": f"예산 목표 조회 실패: {str(e)}"
        }), 500

@app.route('/api/expense-budget-goals', methods=['POST'])
def save_expense_budget_goals():
    """가계부 예산 목표 저장 API"""
    try:
        data = request.get_json()
        user_id = data.get('user_id', 'default')
        expense_goals = data.get('expense_goals', {})  # { "생활": { "전체": 1000000, "외식": 300000 } }
        income_goals = data.get('income_goals', {})    # { "근로소득": { "전체": 5000000, "급여": 5000000 } }

        print(f"Saving expense budget goals: user_id={user_id}")

        result = db_service.save_expense_budget_goals(user_id, expense_goals, income_goals)

        if result.get('status') == 'success':
            return jsonify(result)
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error saving expense budget goals: {e}")
        return jsonify({
            "status": "error",
            "message": f"예산 목표 저장 실패: {str(e)}"
        }), 500

# === 투자 철학 API (MASTER_PLAN Page 1) ===

@app.route('/api/investment-philosophy', methods=['GET'])
def get_investment_philosophy():
    """투자 철학 조회 API"""
    try:
        user_id = request.args.get('user_id', type=int)

        if not user_id:
            return jsonify({
                "status": "error",
                "message": "user_id가 필요합니다."
            }), 400

        result = db_service.get_investment_philosophy(user_id)

        if result.get('status') == 'success':
            return jsonify(result)
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error getting investment philosophy: {e}")
        return jsonify({
            "status": "error",
            "message": f"투자 철학 조회 실패: {str(e)}"
        }), 500

@app.route('/api/investment-philosophy', methods=['POST'])
def save_investment_philosophy():
    """투자 철학 저장 API"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')

        if not user_id:
            return jsonify({
                "status": "error",
                "message": "user_id가 필요합니다."
            }), 400

        philosophy_data = {
            'goal': data.get('goal', {}),
            'forbidden_assets': data.get('forbiddenAssets', []),
            'allocation_range': data.get('allocationRange', []),
            'principles': data.get('principles', []),
            'methods': data.get('methods', [])
        }

        result = db_service.save_investment_philosophy(user_id, philosophy_data)

        if result.get('status') == 'success':
            return jsonify(result)
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error saving investment philosophy: {e}")
        return jsonify({
            "status": "error",
            "message": f"투자 철학 저장 실패: {str(e)}"
        }), 500

# === Page 3: Industries (섹터 & 종목 분석) API ===

@app.route('/api/sector-performance', methods=['GET'])
def get_sector_performance():
    """섹터 성과 조회 API"""
    try:
        user_id = request.args.get('user_id', type=int)
        date = request.args.get('date')

        if not user_id or not date:
            return jsonify({
                "status": "error",
                "message": "user_id와 date는 필수입니다."
            }), 400

        result = db_service.get_sector_performance(user_id, date)

        if result['status'] == 'success':
            return jsonify(result), 200
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error fetching sector performance: {e}")
        return jsonify({
            "status": "error",
            "message": f"섹터 성과 조회 실패: {str(e)}"
        }), 500

@app.route('/api/sector-performance', methods=['POST'])
def save_sector_performance():
    """섹터 성과 저장 API"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        date = data.get('date')
        sectors = data.get('sectors', [])

        if not user_id or not date:
            return jsonify({
                "status": "error",
                "message": "user_id와 date는 필수입니다."
            }), 400

        result = db_service.save_sector_performance(user_id, date, sectors)

        if result['status'] == 'success':
            return jsonify(result), 200
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error saving sector performance: {e}")
        return jsonify({
            "status": "error",
            "message": f"섹터 성과 저장 실패: {str(e)}"
        }), 500

@app.route('/api/watchlist', methods=['GET'])
def get_watchlist():
    """관심 종목 조회 API"""
    try:
        user_id = request.args.get('user_id', type=int)

        if not user_id:
            return jsonify({
                "status": "error",
                "message": "user_id는 필수입니다."
            }), 400

        result = db_service.get_watchlist(user_id)

        if result['status'] == 'success':
            return jsonify(result), 200
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error fetching watchlist: {e}")
        return jsonify({
            "status": "error",
            "message": f"관심 종목 조회 실패: {str(e)}"
        }), 500

@app.route('/api/watchlist', methods=['POST'])
def add_watchlist_item():
    """관심 종목 추가 API"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')

        if not user_id:
            return jsonify({
                "status": "error",
                "message": "user_id는 필수입니다."
            }), 400

        if not data.get('symbol') or not data.get('name'):
            return jsonify({
                "status": "error",
                "message": "symbol과 name은 필수입니다."
            }), 400

        result = db_service.add_watchlist_item(user_id, data)

        if result['status'] == 'success':
            return jsonify(result), 200
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error adding watchlist item: {e}")
        return jsonify({
            "status": "error",
            "message": f"관심 종목 추가 실패: {str(e)}"
        }), 500

@app.route('/api/watchlist/<int:stock_id>', methods=['PUT'])
def update_watchlist_item(stock_id):
    """관심 종목 수정 API"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')

        if not user_id:
            return jsonify({
                "status": "error",
                "message": "user_id는 필수입니다."
            }), 400

        result = db_service.update_watchlist_item(user_id, stock_id, data)

        if result['status'] == 'success':
            return jsonify(result), 200
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error updating watchlist item: {e}")
        return jsonify({
            "status": "error",
            "message": f"관심 종목 수정 실패: {str(e)}"
        }), 500

@app.route('/api/watchlist/<int:stock_id>', methods=['DELETE'])
def delete_watchlist_item(stock_id):
    """관심 종목 삭제 API"""
    try:
        user_id = request.args.get('user_id', type=int)

        if not user_id:
            return jsonify({
                "status": "error",
                "message": "user_id는 필수입니다."
            }), 400

        result = db_service.delete_watchlist_item(user_id, stock_id)

        if result['status'] == 'success':
            return jsonify(result), 200
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error deleting watchlist item: {e}")
        return jsonify({
            "status": "error",
            "message": f"관심 종목 삭제 실패: {str(e)}"
        }), 500

# === 사용자 인증 API ===

@app.route('/api/auth/register', methods=['POST'])
def register():
    """회원가입 API"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '')

        # 입력 검증
        if not username or not password:
            return jsonify({
                "status": "error",
                "message": "사용자명과 비밀번호를 모두 입력해주세요."
            }), 400

        if len(username) < 3:
            return jsonify({
                "status": "error",
                "message": "사용자명은 3글자 이상이어야 합니다."
            }), 400

        if len(password) < 4:
            return jsonify({
                "status": "error",
                "message": "비밀번호는 4글자 이상이어야 합니다."
            }), 400

        result = db_service.create_user(username, password)

        if result.get('status') == 'success':
            # 회원가입 성공 시 JWT 토큰 생성
            user_id = result.get('user_id')
            username = result.get('username')

            if user_id and username:
                token = db_service.generate_jwt_token(user_id, username)
                result['token'] = token

            return jsonify(result), 201
        else:
            return jsonify(result), 400

    except Exception as e:
        print(f"Error registering user: {e}")
        return jsonify({
            "status": "error",
            "message": f"회원가입 실패: {str(e)}"
        }), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    """로그인 API"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '')

        # 입력 검증
        if not username or not password:
            return jsonify({
                "status": "error",
                "message": "사용자명과 비밀번호를 모두 입력해주세요."
            }), 400

        result = db_service.authenticate_user(username, password)

        if result.get('status') == 'success':
            return jsonify(result)
        else:
            return jsonify(result), 401

    except Exception as e:
        print(f"Error logging in user: {e}")
        return jsonify({
            "status": "error",
            "message": f"로그인 실패: {str(e)}"
        }), 500

@app.route('/api/auth/user/<int:user_id>', methods=['GET'])
def get_user(user_id):
    """사용자 정보 조회 API"""
    try:
        user = db_service.get_user_by_id(user_id)

        if user:
            return jsonify({
                "status": "success",
                "user": {
                    "id": user['id'],
                    "username": user['username'],
                    "created_at": user['created_at'].isoformat() if user['created_at'] else None
                }
            })
        else:
            return jsonify({
                "status": "error",
                "message": "사용자를 찾을 수 없습니다."
            }), 404

    except Exception as e:
        print(f"Error getting user: {e}")
        return jsonify({
            "status": "error",
            "message": f"사용자 조회 실패: {str(e)}"
        }), 500

@app.route('/api/admin/delete-user/<username>', methods=['DELETE'])
def admin_delete_user(username):
    """관리자 전용: 사용자 계정 삭제 API"""
    try:
        # 관리자 권한 체크 (임시로 특정 패스워드로 확인)
        admin_password = request.headers.get('X-Admin-Password')
        if admin_password != 'admin-delete-2025':
            return jsonify({
                "status": "error",
                "message": "관리자 권한이 필요합니다."
            }), 403

        result = db_service.delete_user(username)

        if result.get('status') == 'success':
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        print(f"Error deleting user: {e}")
        return jsonify({
            "status": "error",
            "message": f"사용자 삭제 실패: {str(e)}"
        }), 500

@app.route('/api/auth/delete-account', methods=['DELETE'])
@token_required
def delete_own_account(current_user):
    """사용자 자신의 계정 삭제 API"""
    try:
        data = request.get_json()
        password = data.get('password', '')

        if not password:
            return jsonify({
                "status": "error",
                "message": "비밀번호를 입력해주세요."
            }), 400

        # 현재 사용자 정보 확인
        user = db_service.get_user_by_id(current_user['user_id'])
        if not user:
            return jsonify({
                "status": "error",
                "message": "사용자를 찾을 수 없습니다."
            }), 404

        # 비밀번호 확인
        if not db_service._verify_password(password, user['password_hash']):
            return jsonify({
                "status": "error",
                "message": "비밀번호가 일치하지 않습니다."
            }), 401

        # 계정 삭제
        result = db_service.delete_user(user['username'])

        if result.get('status') == 'success':
            return jsonify({
                "status": "success",
                "message": "계정이 성공적으로 삭제되었습니다. 모든 데이터가 영구 삭제됩니다."
            })
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error deleting own account: {e}")
        return jsonify({
            "status": "error",
            "message": f"계정 삭제 실패: {str(e)}"
        }), 500

@app.route('/api/auth/request-password-reset', methods=['POST'])
def request_password_reset():
    """비밀번호 재설정 요청 API"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()

        if not username:
            return jsonify({
                "status": "error",
                "message": "사용자명을 입력해주세요."
            }), 400

        result = db_service.create_password_reset_request(username)

        if result.get('status') == 'success':
            return jsonify({
                "status": "success",
                "message": result.get('message'),
                "reset_token": result.get('reset_token'),  # 실제 운영에서는 이메일로 전송
                "expires_in": result.get('expires_in')
            })
        else:
            return jsonify(result), 400

    except Exception as e:
        print(f"Error requesting password reset: {e}")
        return jsonify({
            "status": "error",
            "message": "비밀번호 재설정 요청 실패"
        }), 500

@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    """비밀번호 재설정 실행 API"""
    try:
        data = request.get_json()
        reset_token = data.get('reset_token', '')
        new_password = data.get('new_password', '')

        if not reset_token or not new_password:
            return jsonify({
                "status": "error",
                "message": "재설정 토큰과 새 비밀번호를 모두 입력해주세요."
            }), 400

        # 토큰 검증
        token_result = db_service.verify_reset_token(reset_token)
        if token_result.get('status') != 'success':
            return jsonify(token_result), 401

        # 비밀번호 업데이트
        user_id = token_result.get('user_id')
        update_result = db_service.update_user_password(user_id, new_password)

        if update_result.get('status') == 'success':
            return jsonify({
                "status": "success",
                "message": "비밀번호가 성공적으로 재설정되었습니다. 새 비밀번호로 로그인해주세요."
            })
        else:
            return jsonify(update_result), 400

    except Exception as e:
        print(f"Error resetting password: {e}")
        return jsonify({
            "status": "error",
            "message": "비밀번호 재설정 실패"
        }), 500

@app.route('/api/auth/change-password', methods=['PUT'])
@token_required
def change_password(current_user):
    """로그인된 사용자의 비밀번호 변경 API"""
    try:
        data = request.get_json()
        current_password = data.get('current_password', '')
        new_password = data.get('new_password', '')

        if not current_password or not new_password:
            return jsonify({
                "status": "error",
                "message": "현재 비밀번호와 새 비밀번호를 모두 입력해주세요."
            }), 400

        # 현재 사용자 정보 확인
        user = db_service.get_user_by_id(current_user['user_id'])
        if not user:
            return jsonify({
                "status": "error",
                "message": "사용자를 찾을 수 없습니다."
            }), 404

        # 현재 비밀번호 확인
        if not db_service._verify_password(current_password, user['password_hash']):
            return jsonify({
                "status": "error",
                "message": "현재 비밀번호가 일치하지 않습니다."
            }), 401

        # 새 비밀번호로 업데이트
        update_result = db_service.update_user_password(current_user['user_id'], new_password)

        if update_result.get('status') == 'success':
            return jsonify({
                "status": "success",
                "message": "비밀번호가 성공적으로 변경되었습니다."
            })
        else:
            return jsonify(update_result), 400

    except Exception as e:
        print(f"Error changing password: {e}")
        print(f"Error type: {type(e)}")
        print(f"Current user data: {current_user}")
        return jsonify({
            "status": "error",
            "message": f"비밀번호 변경 실패: {str(e)}"
        }), 500

@app.route('/api/debug/user/<username>')
def debug_user_hash(username):
    """임시 디버그: 사용자 해시 정보 확인"""
    try:
        user = db_service.get_user_by_username(username)
        if user:
            hash_format = "unknown"
            if user['password_hash'].startswith('$2'):
                hash_format = "bcrypt"
            elif ':' in user['password_hash']:
                hash_format = "pbkdf2"

            return jsonify({
                "username": user['username'],
                "hash_format": hash_format,
                "hash_prefix": user['password_hash'][:30] + "...",
                "created_at": str(user.get('created_at'))
            })
        else:
            return jsonify({"error": "User not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/debug/users')
def debug_all_users():
    """임시 디버그: 모든 사용자 목록 조회"""
    try:
        with db_service.get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id, username, created_at FROM users ORDER BY created_at DESC")
                users = cur.fetchall()

                user_list = []
                for user in users:
                    user_list.append({
                        "id": user['id'],
                        "username": user['username'],
                        "created_at": str(user['created_at'])
                    })

                return jsonify({
                    "total_users": len(user_list),
                    "users": user_list
                })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/portfolio-history', methods=['GET'])
def get_portfolio_history():
    """포트폴리오 이력 조회 API"""
    try:
        user_id = request.args.get('user_id')
        if not user_id:
            return jsonify({"status": "error", "message": "user_id required"}), 400

        time_range = request.args.get('time_range', 'daily')  # annual, monthly, daily
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        result = db_service.get_portfolio_history(user_id, time_range, start_date, end_date)
        return jsonify(result)

    except Exception as e:
        print(f"Error getting portfolio history: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/portfolio-history/summary', methods=['POST'])
def create_daily_summary():
    """일일 포트폴리오 요약 생성 API"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')

        if not user_id:
            return jsonify({"status": "error", "message": "user_id required"}), 400

        result = db_service.create_daily_summary(user_id)
        return jsonify(result)

    except Exception as e:
        print(f"Error creating daily summary: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

# ========== 가계부/지출관리 API 엔드포인트 ==========

@app.route('/api/expenses', methods=['POST'])
def add_expense():
    """거래내역 추가 API (포트폴리오 add_asset 패턴과 동일)"""
    try:
        expense_data = request.get_json()

        # 필수 필드 검증 (user_id 포함)
        required_fields = ['user_id', 'transaction_type', 'amount', 'category', 'subcategory', 'transaction_date']
        for field in required_fields:
            if field not in expense_data or not expense_data[field]:
                return jsonify({
                    "status": "error",
                    "message": f"필수 필드가 누락되었습니다: {field}"
                }), 400

        user_id = expense_data.get('user_id')

        # 거래내역 저장
        result = db_service.add_expense(user_id, expense_data)

        if result["status"] == "success":
            return jsonify(result), 201
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error in add_expense: {e}")
        return jsonify({
            "status": "error",
            "message": f"거래내역 추가 중 오류가 발생했습니다: {str(e)}"
        }), 500

@app.route('/api/expenses', methods=['GET'])
def get_expenses():
    """거래내역 조회 API (필터링 지원)"""
    try:
        # 쿼리 파라미터에서 user_id 가져오기 (포트폴리오 패턴과 동일)
        user_id = request.args.get('user_id')

        if user_id:
            user_id = int(user_id)
        else:
            # user_id가 없으면 기본값 사용 (포트폴리오와 동일한 패턴)
            user_id = 1

        # 쿼리 파라미터에서 필터링 옵션 추출
        filters = {}
        if request.args.get('start_date'):
            filters['start_date'] = request.args.get('start_date')
        if request.args.get('end_date'):
            filters['end_date'] = request.args.get('end_date')
        if request.args.get('category'):
            filters['category'] = request.args.get('category')
        if request.args.get('transaction_type'):
            filters['transaction_type'] = request.args.get('transaction_type')

        result = db_service.get_all_expenses(user_id, filters)
        return jsonify(result)

    except Exception as e:
        print(f"Error in get_expenses: {e}")
        return jsonify({
            "status": "error",
            "message": f"거래내역 조회 중 오류가 발생했습니다: {str(e)}"
        }), 500


@app.route('/api/expenses/yearly', methods=['GET'])
def get_expenses_yearly():
    """연간 월별 지출/수입 합계 조회 API"""
    try:
        user_id = request.args.get('user_id')
        year = request.args.get('year')

        if user_id:
            user_id = int(user_id)
        else:
            user_id = 1

        if not year:
            from datetime import datetime
            year = datetime.now().year
        else:
            year = int(year)

        # 해당 연도의 전체 데이터 조회
        filters = {
            'start_date': f'{year}-01-01',
            'end_date': f'{year}-12-31'
        }

        result = db_service.get_all_expenses(user_id, filters)

        if result.get('status') != 'success':
            return jsonify(result), 500

        # 월별로 집계
        monthly_data = {}
        for month in range(1, 13):
            monthly_data[month] = {'expense': 0, 'income': 0, 'transfer': 0}

        for expense in result.get('expenses', []):
            date_str = expense.get('date', '')
            if date_str:
                try:
                    month = int(date_str.split('-')[1])
                    amount = float(expense.get('amount', 0))
                    tx_type = expense.get('transaction_type', '지출')

                    if tx_type == '지출':
                        monthly_data[month]['expense'] += amount
                    elif tx_type == '수입':
                        monthly_data[month]['income'] += amount
                    elif tx_type == '이체':
                        monthly_data[month]['transfer'] += amount
                except (IndexError, ValueError):
                    continue

        # 응답 형식으로 변환
        chart_data = []
        for month in range(1, 13):
            chart_data.append({
                'month': f'{month}월',
                'expense': monthly_data[month]['expense'],
                'income': monthly_data[month]['income'],
                'transfer': monthly_data[month]['transfer'],
                'net': monthly_data[month]['income'] - monthly_data[month]['expense']
            })

        return jsonify({
            'status': 'success',
            'year': year,
            'data': chart_data
        })

    except Exception as e:
        print(f"Error in get_expenses_yearly: {e}")
        return jsonify({
            "status": "error",
            "message": f"연간 데이터 조회 중 오류가 발생했습니다: {str(e)}"
        }), 500


@app.route('/api/expenses/export/excel', methods=['GET'])
def export_expenses_excel():
    """가계부 데이터를 Excel 파일로 다운로드"""
    try:
        # 쿼리 파라미터에서 user_id, 연도/월 가져오기
        user_id = request.args.get('user_id')
        if not user_id:
            return jsonify({"status": "error", "message": "user_id required"}), 400

        user_id = int(user_id)

        # 날짜 필터 (선택적)
        filters = {}
        year = request.args.get('year')
        month = request.args.get('month')

        if year and month:
            filters['start_date'] = f"{year}-{month.zfill(2)}-01"
            # 월의 마지막 날 계산
            import calendar
            last_day = calendar.monthrange(int(year), int(month))[1]
            filters['end_date'] = f"{year}-{month.zfill(2)}-{last_day}"

        # PostgreSQL에서 사용자별 거래내역 데이터 조회
        result = db_service.get_all_expenses(user_id, filters)

        if result.get('status') != 'success':
            return jsonify({"status": "error", "message": "데이터 조회 실패"}), 500

        expenses = result.get('data', [])

        # Excel 파일 생성
        wb = Workbook()

        # Sheet 1: 거래내역
        ws1 = wb.active
        ws1.title = "거래내역"

        # 헤더 스타일
        header_fill = PatternFill(start_color="50C878", end_color="50C878", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # 헤더
        headers = ['날짜', '거래유형', '대분류', '소분류', '거래처', '금액', '메모']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 데이터
        for row_idx, expense in enumerate(expenses, 2):
            ws1.cell(row=row_idx, column=1, value=expense.get('transaction_date', ''))
            ws1.cell(row=row_idx, column=2, value=expense.get('transaction_type', ''))
            ws1.cell(row=row_idx, column=3, value=expense.get('main_category', ''))
            ws1.cell(row=row_idx, column=4, value=expense.get('sub_category', ''))
            ws1.cell(row=row_idx, column=5, value=expense.get('merchant', ''))
            ws1.cell(row=row_idx, column=6, value=expense.get('amount', 0))
            ws1.cell(row=row_idx, column=7, value=expense.get('memo', ''))

        # 컬럼 너비 조정
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 10
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 15
        ws1.column_dimensions['E'].width = 20
        ws1.column_dimensions['F'].width = 12
        ws1.column_dimensions['G'].width = 30

        # Sheet 2: 카테고리별 요약
        ws2 = wb.create_sheet("카테고리별 요약")

        # 헤더
        headers2 = ['대분류', '소분류', '총액', '건수', '비중(%)']
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 카테고리별 집계
        category_summary = {}
        total_amount = sum(expense.get('amount', 0) for expense in expenses)

        for expense in expenses:
            main_cat = expense.get('main_category', '기타')
            sub_cat = expense.get('sub_category', '기타')
            amount = expense.get('amount', 0)

            key = f"{main_cat}_{sub_cat}"

            if key not in category_summary:
                category_summary[key] = {
                    'main_category': main_cat,
                    'sub_category': sub_cat,
                    'total': 0,
                    'count': 0
                }

            category_summary[key]['total'] += amount
            category_summary[key]['count'] += 1

        # 데이터
        for row_idx, summary in enumerate(category_summary.values(), 2):
            ws2.cell(row=row_idx, column=1, value=summary['main_category'])
            ws2.cell(row=row_idx, column=2, value=summary['sub_category'])
            ws2.cell(row=row_idx, column=3, value=summary['total'])
            ws2.cell(row=row_idx, column=4, value=summary['count'])
            ws2.cell(row=row_idx, column=5, value=round((summary['total'] / total_amount * 100) if total_amount > 0 else 0, 2))

        # 컬럼 너비 조정
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws2.column_dimensions[col].width = 15

        # Sheet 3: 월간 요약
        ws3 = wb.create_sheet("월간 요약")

        # 헤더
        headers3 = ['항목', '값']
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 지출/수입 집계
        total_expense = sum(expense.get('amount', 0) for expense in expenses if expense.get('transaction_type') == '지출')
        total_income = sum(expense.get('amount', 0) for expense in expenses if expense.get('transaction_type') == '수입')
        net_change = total_income - total_expense

        ws3.cell(row=2, column=1, value="총 지출")
        ws3.cell(row=2, column=2, value=total_expense)
        ws3.cell(row=3, column=1, value="총 수입")
        ws3.cell(row=3, column=2, value=total_income)
        ws3.cell(row=4, column=1, value="순자산 변화")
        ws3.cell(row=4, column=2, value=net_change)
        ws3.cell(row=5, column=1, value="총 거래 건수")
        ws3.cell(row=5, column=2, value=len(expenses))

        # 컬럼 너비 조정
        ws3.column_dimensions['A'].width = 15
        ws3.column_dimensions['B'].width = 20

        # 파일명 생성
        if year and month:
            filename = f"expenses_{year}_{month.zfill(2)}.xlsx"
        else:
            filename = f"expenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # 메모리 버퍼에 저장
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error exporting expenses to Excel: {e}")
        return jsonify({
            "status": "error",
            "message": f"Excel 생성 실패: {str(e)}"
        }), 500

@app.route('/api/expenses/<int:expense_id>', methods=['PUT'])
def update_expense(expense_id):
    """거래내역 수정 API"""
    try:
        expense_data = request.get_json()

        # user_id를 body에서 가져오기 (포트폴리오 패턴과 동일)
        user_id = expense_data.get('user_id')
        if not user_id:
            return jsonify({"status": "error", "message": "user_id가 필요합니다."}), 400

        result = db_service.update_expense(expense_id, expense_data, user_id)

        if result["status"] == "success":
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        print(f"Error in update_expense: {e}")
        return jsonify({
            "status": "error",
            "message": f"거래내역 수정 중 오류가 발생했습니다: {str(e)}"
        }), 500

@app.route('/api/expenses/<int:expense_id>', methods=['DELETE'])
def delete_expense(expense_id):
    """거래내역 삭제 API"""
    try:
        # user_id를 쿼리 파라미터에서 가져오기 (포트폴리오 패턴과 동일)
        user_id = request.args.get('user_id')
        if not user_id:
            return jsonify({"status": "error", "message": "user_id가 필요합니다."}), 400

        user_id = int(user_id)
        result = db_service.delete_expense(expense_id, user_id)

        if result["status"] == "success":
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        print(f"Error in delete_expense: {e}")
        return jsonify({
            "status": "error",
            "message": f"거래내역 삭제 중 오류가 발생했습니다: {str(e)}"
        }), 500

@app.route('/api/budgets', methods=['POST'])
def set_budget():
    """예산 설정 API"""
    try:
        # JWT 토큰에서 사용자 확인
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({"status": "error", "message": "인증이 필요합니다."}), 401

        try:
            token = auth_header.split(' ')[1]
            user_data = db_service.verify_jwt_token(token)
            if not user_data:
                return jsonify({"status": "error", "message": "유효하지 않은 토큰입니다."}), 401
            user_id = user_data['user_id']
        except:
            return jsonify({"status": "error", "message": "토큰 검증에 실패했습니다."}), 401

        budget_data = request.get_json()

        # 필수 필드 검증
        required_fields = ['category', 'monthly_budget', 'year', 'month']
        for field in required_fields:
            if field not in budget_data:
                return jsonify({
                    "status": "error",
                    "message": f"필수 필드가 누락되었습니다: {field}"
                }), 400

        result = db_service.set_budget(user_id, budget_data)

        if result["status"] == "success":
            return jsonify(result), 201
        else:
            return jsonify(result), 500

    except Exception as e:
        print(f"Error in set_budget: {e}")
        return jsonify({
            "status": "error",
            "message": f"예산 설정 중 오류가 발생했습니다: {str(e)}"
        }), 500

@app.route('/api/budgets', methods=['GET'])
def get_budgets():
    """예산 조회 API"""
    try:
        # JWT 토큰에서 사용자 확인
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({"status": "error", "message": "인증이 필요합니다."}), 401

        try:
            token = auth_header.split(' ')[1]
            user_data = db_service.verify_jwt_token(token)
            if not user_data:
                return jsonify({"status": "error", "message": "유효하지 않은 토큰입니다."}), 401
            user_id = user_data['user_id']
        except:
            return jsonify({"status": "error", "message": "토큰 검증에 실패했습니다."}), 401

        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)

        result = db_service.get_budgets(user_id, year, month)
        return jsonify(result)

    except Exception as e:
        print(f"Error in get_budgets: {e}")
        return jsonify({
            "status": "error",
            "message": f"예산 조회 중 오류가 발생했습니다: {str(e)}"
        }), 500

@app.route('/api/budgets/progress', methods=['GET'])
def get_budget_progress():
    """예산 진행률 조회 API"""
    try:
        # JWT 토큰에서 사용자 확인
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({"status": "error", "message": "인증이 필요합니다."}), 401

        try:
            token = auth_header.split(' ')[1]
            user_data = db_service.verify_jwt_token(token)
            if not user_data:
                return jsonify({"status": "error", "message": "유효하지 않은 토큰입니다."}), 401
            user_id = user_data['user_id']
        except:
            return jsonify({"status": "error", "message": "토큰 검증에 실패했습니다."}), 401

        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)

        if not year or not month:
            return jsonify({
                "status": "error",
                "message": "year와 month 파라미터가 필요합니다."
            }), 400

        result = db_service.get_budget_progress(user_id, year, month)
        return jsonify(result)

    except Exception as e:
        print(f"Error in get_budget_progress: {e}")
        return jsonify({
            "status": "error",
            "message": f"예산 진행률 조회 중 오류가 발생했습니다: {str(e)}"
        }), 500


@app.route('/api/debug/cleanup-users', methods=['POST'])
def cleanup_all_users():
    """임시 디버그: 모든 사용자 계정 삭제"""
    try:
        with db_service.get_connection() as conn:
            with conn.cursor() as cur:
                # 모든 사용자 삭제 (CASCADE로 관련 데이터도 삭제)
                cur.execute("DELETE FROM users")
                deleted_count = cur.rowcount

                return jsonify({
                    "status": "success",
                    "message": f"{deleted_count}개의 사용자 계정이 삭제되었습니다.",
                    "deleted_count": deleted_count
                })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# 금리지표 API 엔드포인트들
@app.route('/api/rawdata/federal-funds-rate')
def get_federal_funds_rate_rawdata():
    """Get Federal Funds Rate raw data"""
    try:
        data = get_federal_funds_rate()

        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Federal Funds Rate"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/federal-funds-rate')
def get_federal_funds_rate_history():
    try:
        url = "https://www.investing.com/economic-calendar/federal-funds-rate-169"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({
                "status": "success",
                "data": history_data,
                "source": "investing.com"
            })
        else:
            return jsonify({
                "status": "error",
                "message": "Failed to fetch history data"
            }), 500
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/rawdata/core-cpi')
def get_core_cpi_rawdata():
    """Get Core CPI raw data"""
    try:
        data = get_core_cpi()

        if "error" in data:
            return jsonify({
                "status": "error",
                "message": data["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": data,
            "source": "investing.com",
            "indicator": "Core CPI"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/core-cpi')
def get_core_cpi_history():
    try:
        url = "https://www.investing.com/economic-calendar/core-cpi-56"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({
                "status": "success",
                "data": history_data,
                "source": "investing.com"
            })
        else:
            return jsonify({
                "status": "error",
                "message": "Failed to fetch history data"
            }), 500
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/rawdata/ten-year-treasury')
def get_ten_year_treasury_rawdata():
    """Get 10-Year Treasury raw data"""
    try:
        result = CrawlerService.crawl_indicator('ten-year-treasury')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": {
                "latest_release": result["latest_release"],
                "next_release": result.get("next_release")
            },
            "source": "rates-bonds",
            "indicator": "10-Year Treasury"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/ten-year-treasury')
def get_ten_year_treasury_history():
    try:
        result = CrawlerService.crawl_indicator('ten-year-treasury')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": result["history_table"],
            "source": "rates-bonds"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/rawdata/two-year-treasury')
def get_two_year_treasury_rawdata():
    """Get 2-Year Treasury raw data"""
    try:
        result = CrawlerService.crawl_indicator('two-year-treasury')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": {
                "latest_release": result["latest_release"],
                "next_release": result.get("next_release")
            },
            "source": "rates-bonds",
            "indicator": "2-Year Treasury"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/two-year-treasury')
def get_two_year_treasury_history():
    try:
        result = CrawlerService.crawl_indicator('two-year-treasury')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": result["history_table"],
            "source": "rates-bonds"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/rawdata/yield-curve-10y-2y')
def get_yield_curve_rawdata():
    """Get Yield Curve (10Y-2Y) raw data"""
    try:
        result = CrawlerService.crawl_indicator('yield-curve-10y-2y')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": {
                "latest_release": result["latest_release"],
                "next_release": result.get("next_release")
            },
            "source": "fred",
            "indicator": "Yield Curve (10Y-2Y)"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/yield-curve-10y-2y')
def get_yield_curve_history():
    try:
        result = CrawlerService.crawl_indicator('yield-curve-10y-2y')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": result["history_table"],
            "source": "fred"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

# =============================================================================
# Phase 2 Sentiment Indicators API
# =============================================================================

@app.route('/api/rawdata/sp500-pe')
def get_sp500_pe_rawdata():
    """Get S&P 500 P/E Ratio raw data"""
    try:
        result = CrawlerService.crawl_indicator('sp500-pe')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        # 데이터베이스에 저장
        db_service.save_indicator_data('sp500-pe', result)

        return jsonify({
            "status": "success",
            "data": {
                "latest_release": result.get("latest_release"),
                "next_release": result.get("next_release")
            },
            "source": "multpl.com",
            "indicator": "S&P 500 P/E Ratio"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/sp500-pe')
def get_sp500_pe_history():
    try:
        result = CrawlerService.crawl_indicator('sp500-pe')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": result.get("history", []),
            "source": "multpl.com"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/rawdata/shiller-pe')
def get_shiller_pe_rawdata():
    """Get Shiller P/E Ratio (CAPE) raw data"""
    try:
        result = CrawlerService.crawl_indicator('shiller-pe')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        # 데이터베이스에 저장
        db_service.save_indicator_data('shiller-pe', result)

        return jsonify({
            "status": "success",
            "data": {
                "latest_release": result.get("latest_release"),
                "next_release": result.get("next_release")
            },
            "source": "multpl.com",
            "indicator": "Shiller P/E Ratio (CAPE)"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/shiller-pe')
def get_shiller_pe_history():
    try:
        result = CrawlerService.crawl_indicator('shiller-pe')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": result.get("history", []),
            "source": "multpl.com"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/rawdata/put-call-ratio')
def get_put_call_ratio_rawdata():
    """Get CBOE Put/Call Ratio raw data"""
    try:
        result = CrawlerService.crawl_indicator('put-call-ratio')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        # 데이터베이스에 저장
        db_service.save_indicator_data('put-call-ratio', result)

        return jsonify({
            "status": "success",
            "data": {
                "latest_release": result.get("latest_release"),
                "next_release": result.get("next_release")
            },
            "source": "cboe (fallback)",
            "indicator": "CBOE Put/Call Ratio",
            "note": result.get("note", "")
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/put-call-ratio')
def get_put_call_ratio_history():
    try:
        result = CrawlerService.crawl_indicator('put-call-ratio')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": result.get("history", []),
            "source": "cboe (fallback)",
            "note": result.get("note", "")
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/rawdata/real-yield-tips')
def get_real_yield_rawdata():
    """Get Real Yield (TIPS) raw data"""
    try:
        result = CrawlerService.crawl_indicator('real-yield-tips')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": {
                "latest_release": result["latest_release"],
                "next_release": result.get("next_release")
            },
            "source": "fred",
            "indicator": "Real Yield (TIPS)"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/history-table/real-yield-tips')
def get_real_yield_history():
    try:
        result = CrawlerService.crawl_indicator('real-yield-tips')

        if "error" in result:
            return jsonify({
                "status": "error",
                "message": result["error"]
            }), 500

        return jsonify({
            "status": "success",
            "data": result["history_table"],
            "source": "fred"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

# 무역지표 API 엔드포인트들
@app.route('/api/rawdata/trade-balance')
def get_trade_balance_rawdata():
    """Get Trade Balance raw data"""
    try:
        data = get_trade_balance()
        if "error" in data:
            return jsonify({"status": "error", "message": data["error"]}), 500
        return jsonify({"status": "success", "data": data, "source": "investing.com", "indicator": "Trade Balance"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/history-table/trade-balance')
def get_trade_balance_history():
    try:
        url = "https://www.investing.com/economic-calendar/trade-balance-602"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({"status": "success", "data": history_data, "source": "investing.com"})
        else:
            return jsonify({"status": "error", "message": "Failed to fetch history data"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/rawdata/exports')
def get_exports_rawdata():
    """Get Exports raw data"""
    try:
        data = get_exports()
        if "error" in data:
            return jsonify({"status": "error", "message": data["error"]}), 500
        return jsonify({"status": "success", "data": data, "source": "investing.com", "indicator": "Exports"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/history-table/exports')
def get_exports_history():
    try:
        url = "https://www.investing.com/economic-calendar/exports-605"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({"status": "success", "data": history_data, "source": "investing.com"})
        else:
            return jsonify({"status": "error", "message": "Failed to fetch history data"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/rawdata/imports')
def get_imports_rawdata():
    """Get Imports raw data"""
    try:
        data = get_imports()
        if "error" in data:
            return jsonify({"status": "error", "message": data["error"]}), 500
        return jsonify({"status": "success", "data": data, "source": "investing.com", "indicator": "Imports"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/history-table/imports')
def get_imports_history():
    try:
        url = "https://www.investing.com/economic-calendar/imports-604"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({"status": "success", "data": history_data, "source": "investing.com"})
        else:
            return jsonify({"status": "error", "message": "Failed to fetch history data"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/rawdata/current-account')
def get_current_account_rawdata():
    """Get Current Account raw data"""
    try:
        data = get_current_account()
        if "error" in data:
            return jsonify({"status": "error", "message": data["error"]}), 500
        return jsonify({"status": "success", "data": data, "source": "investing.com", "indicator": "Current Account"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/history-table/current-account')
def get_current_account_history():
    try:
        url = "https://www.investing.com/economic-calendar/current-account-603"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({"status": "success", "data": history_data, "source": "investing.com"})
        else:
            return jsonify({"status": "error", "message": "Failed to fetch history data"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

# 물가지표 API 엔드포인트들
@app.route('/api/rawdata/cpi')
def get_cpi_rawdata():
    """Get CPI raw data"""
    try:
        data = get_cpi()
        if "error" in data:
            return jsonify({"status": "error", "message": data["error"]}), 500
        return jsonify({"status": "success", "data": data, "source": "investing.com", "indicator": "CPI"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/history-table/cpi')
def get_cpi_history():
    try:
        url = "https://www.investing.com/economic-calendar/cpi-69"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({"status": "success", "data": history_data, "source": "investing.com"})
        else:
            return jsonify({"status": "error", "message": "Failed to fetch history data"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/rawdata/ppi')
def get_ppi_rawdata():
    """Get PPI raw data"""
    try:
        data = get_ppi()
        if "error" in data:
            return jsonify({"status": "error", "message": data["error"]}), 500
        return jsonify({"status": "success", "data": data, "source": "investing.com", "indicator": "PPI"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/history-table/ppi')
def get_ppi_history():
    try:
        url = "https://www.investing.com/economic-calendar/ppi-238"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({"status": "success", "data": history_data, "source": "investing.com"})
        else:
            return jsonify({"status": "error", "message": "Failed to fetch history data"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/rawdata/pce')
def get_pce_rawdata():
    """Get PCE raw data"""
    try:
        data = get_pce()
        if "error" in data:
            return jsonify({"status": "error", "message": data["error"]}), 500
        return jsonify({"status": "success", "data": data, "source": "investing.com", "indicator": "PCE"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/history-table/pce')
def get_pce_history():
    try:
        url = "https://www.investing.com/economic-calendar/pce-price-index-905"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({"status": "success", "data": history_data, "source": "investing.com"})
        else:
            return jsonify({"status": "error", "message": "Failed to fetch history data"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

# 정책지표 API 엔드포인트들
@app.route('/api/rawdata/fomc-minutes')
def get_fomc_minutes_rawdata():
    """Get FOMC Minutes raw data"""
    try:
        data = get_fomc_minutes()
        if "error" in data:
            return jsonify({"status": "error", "message": data["error"]}), 500
        return jsonify({"status": "success", "data": data, "source": "investing.com", "indicator": "FOMC Minutes"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/history-table/fomc-minutes')
def get_fomc_minutes_history():
    try:
        url = "https://www.investing.com/economic-calendar/fomc-meeting-minutes-108"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({"status": "success", "data": history_data, "source": "investing.com"})
        else:
            return jsonify({"status": "error", "message": "Failed to fetch history data"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/rawdata/consumer-confidence')
def get_consumer_confidence_rawdata():
    """Get Consumer Confidence raw data"""
    try:
        data = get_consumer_confidence()
        if "error" in data:
            return jsonify({"status": "error", "message": data["error"]}), 500
        return jsonify({"status": "success", "data": data, "source": "investing.com", "indicator": "Consumer Confidence"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/history-table/consumer-confidence')
def get_consumer_confidence_history():
    try:
        url = "https://www.investing.com/economic-calendar/consumer-confidence-48"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({"status": "success", "data": history_data, "source": "investing.com"})
        else:
            return jsonify({"status": "error", "message": "Failed to fetch history data"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/rawdata/business-inventories')
def get_business_inventories_rawdata():
    """Get Business Inventories raw data"""
    try:
        data = get_business_inventories()
        if "error" in data:
            return jsonify({"status": "error", "message": data["error"]}), 500
        return jsonify({"status": "success", "data": data, "source": "investing.com", "indicator": "Business Inventories"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/history-table/business-inventories')
def get_business_inventories_history():
    try:
        url = "https://www.investing.com/economic-calendar/business-inventories-235"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({"status": "success", "data": history_data, "source": "investing.com"})
        else:
            return jsonify({"status": "error", "message": "Failed to fetch history data"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/rawdata/leading-indicators')
def get_leading_indicators_rawdata():
    """Get Leading Indicators raw data"""
    try:
        data = get_leading_indicators()
        if "error" in data:
            return jsonify({"status": "error", "message": data["error"]}), 500
        return jsonify({"status": "success", "data": data, "source": "investing.com", "indicator": "Leading Indicators"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/history-table/leading-indicators')
def get_leading_indicators_history():
    try:
        url = "https://www.investing.com/economic-calendar/us-leading-index-1968"
        html_content = fetch_html(url)
        if html_content:
            history_data = parse_history_table(html_content)
            return jsonify({"status": "success", "data": history_data, "source": "investing.com"})
        else:
            return jsonify({"status": "error", "message": "Failed to fetch history data"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"Internal server error: {str(e)}"}), 500

# ==================== 코인 거래 분석 API ====================
# Note: 코인 관련 기능은 개발 중이며, 필요한 모듈이 없으면 비활성화됩니다.
try:
    from services.upbit_service import UpbitService
    from services.crypto_analyzer import CryptoTradeAnalyzer
    CRYPTO_ENABLED = True
except ImportError:
    print("Warning: Crypto services not available. Crypto endpoints will be disabled.")
    CRYPTO_ENABLED = False

if CRYPTO_ENABLED:
    # 업비트 API 키 (환경변수에서 로드)
    UPBIT_ACCESS_KEY = os.getenv('UPBIT_ACCESS_KEY', 'Z3mZzfH1Bn61JqqenCyL77tnsvB7jSHQESAAbwN5')
    UPBIT_SECRET_KEY = os.getenv('UPBIT_SECRET_KEY', 'G1INpa7Ac1ewldqkAJLbG9hyUu2CEZIWFADJ9jc9')

    @app.route('/api/crypto/analysis')
    def get_crypto_analysis():
        """코인 거래 분석 결과 조회"""
        try:
            max_orders = request.args.get('max_orders', 300, type=int)

            # 업비트 API 호출
            upbit = UpbitService(UPBIT_ACCESS_KEY, UPBIT_SECRET_KEY)
            orders = upbit.get_all_trades(max_orders=max_orders)

            # 코인별 분류
            coin_trades = {}
            for order in orders:
                market = order.get('market', '')
                if not market.startswith('KRW-'):
                    continue

                coin = market.replace('KRW-', '')
                side = order.get('side')

                if coin not in coin_trades:
                    coin_trades[coin] = {'buys': [], 'sells': []}

                dt = datetime.fromisoformat(order.get('created_at').replace('+09:00', ''))

                # 거래 금액 계산
                executed_funds = order.get('executed_funds')
                if executed_funds is None:
                    volume = float(order.get('executed_volume', 0))
                    avg_price = float(order.get('avg_price', 0)) if order.get('avg_price') else 0
                    executed_funds = volume * avg_price
                else:
                    executed_funds = float(executed_funds)

                paid_fee = float(order.get('paid_fee', 0))

                if side == 'bid':
                    total_amount = executed_funds + paid_fee
                else:
                    total_amount = executed_funds - paid_fee

                trade_info = {
                    'date': dt.strftime('%m.%d %H:%M'),
                    'volume': float(order.get('executed_volume', 0)),
                    'price': float(order.get('avg_price', 0)) if order.get('avg_price') else 0,
                    'total': total_amount,
                    'fee': paid_fee,
                    'side': side
                }

                if side == 'bid':
                    coin_trades[coin]['buys'].append(trade_info)
                elif side == 'ask':
                    coin_trades[coin]['sells'].append(trade_info)

            # 각 코인별 분석
            analyzer = CryptoTradeAnalyzer()
            results = {}

            for coin, trades in coin_trades.items():
                if not trades['buys'] or not trades['sells']:
                    continue

                # 같은 시간대 거래 합산
                buys_grouped = analyzer.group_by_same_time(trades['buys'])
                sells_grouped = analyzer.group_by_same_time(trades['sells'])

                # 라운드 매칭
                rounds = analyzer.match_buy_sell_rounds(buys_grouped, sells_grouped)

                results[coin] = {
                    'rounds': rounds,
                    'total_profit': sum(r['profit'] for r in rounds),
                    'total_rounds': len(rounds)
                }

            return jsonify({
                "status": "success",
                "data": results,
                "total_orders": len(orders)
            })

        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route('/api/crypto/balance')
    def get_crypto_balance():
        """현재 코인 잔고 조회"""
        try:
            upbit = UpbitService(UPBIT_ACCESS_KEY, UPBIT_SECRET_KEY)
            accounts = upbit.get_accounts()

            # KRW를 제외한 코인만 필터링
            balances = []
            for acc in accounts:
                currency = acc.get('currency')
                if currency == 'KRW':
                    continue

                balance = float(acc.get('balance', 0))
                if balance > 0:
                    balances.append({
                        'currency': currency,
                        'balance': balance,
                        'avg_buy_price': float(acc.get('avg_buy_price', 0)),
                        'locked': float(acc.get('locked', 0))
                    })

            return jsonify({
                "status": "success",
                "data": balances
            })

        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500
else:
    # Crypto disabled - provide placeholder endpoints
    @app.route('/api/crypto/analysis')
    def get_crypto_analysis():
        return jsonify({"status": "error", "message": "Crypto services are currently disabled"}), 503

    @app.route('/api/crypto/balance')
    def get_crypto_balance():
        return jsonify({"status": "error", "message": "Crypto services are currently disabled"}), 503

# ========== 통합 경제지표 API (신규) ==========

@app.route('/api/v3/indicators/<indicator_id>')
def get_indicator_v3(indicator_id):
    """
    통합 지표 API - unified_crawler 사용

    예시: /api/v3/indicators/ism-manufacturing
          /api/v3/indicators/cpi
          /api/v3/indicators/unemployment-rate
    """
    try:
        result = crawl_indicator(indicator_id)

        if result["status"] == "error":
            return jsonify(result), 404 if "Unknown indicator" in result.get("message", "") else 500

        return jsonify(result)

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/v3/indicators/category/<category>')
def get_category_indicators_v3(category):
    """
    카테고리별 지표 크롤링

    예시: /api/v3/indicators/category/business
          /api/v3/indicators/category/employment
    """
    try:
        result = crawl_category(category)

        if result.get("status") == "error":
            return jsonify(result), 404

        return jsonify(result)

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

@app.route('/api/v3/indicators')
def get_all_indicators_metadata():
    """
    모든 지표의 메타데이터 반환 (크롤링 없음, 빠름)

    프론트엔드가 지표 목록을 알기 위해 사용
    """
    try:
        enabled_indicators = get_all_enabled_indicators()

        metadata = {
            "total": len(enabled_indicators),
            "categories": CATEGORIES,
            "indicators": {
                id: {
                    "id": config.id,
                    "name": config.name,
                    "name_ko": config.name_ko,
                    "category": config.category,
                    "threshold": config.threshold,
                    "reverse_color": config.reverse_color
                }
                for id, config in enabled_indicators.items()
            }
        }

        return jsonify({
            "status": "success",
            "data": metadata
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Internal server error: {str(e)}"
        }), 500

# ========== 경제 사이클 API ==========

@app.route('/api/v2/macro-cycle', methods=['GET'])
def get_macro_cycle():
    """
    거시경제 사이클 계산 및 국면 판별

    Returns:
        {
            "status": "success",
            "data": {
                "score": 0-100,
                "phase": "침체|회복|확장|둔화",
                "phase_en": "Recession|Early Expansion|...",
                "color": "red|green|emerald|amber",
                "description": "현재 국면 설명",
                "action": "투자 행동 추천",
                "confidence": 0-100,
                "indicators": {...},
                "last_updated": "ISO timestamp"
            }
        }
    """
    try:
        # MacroCycleService 초기화
        cycle_service = MacroCycleService(db_service)

        # 사이클 계산
        result = cycle_service.calculate_cycle()

        if result.get('score') is None:
            return jsonify({
                "status": "error",
                "message": result.get('description', '데이터를 불러올 수 없습니다')
            }), 503

        return jsonify({
            "status": "success",
            "data": result
        })

    except Exception as e:
        import traceback
        print(f"Error in get_macro_cycle: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"사이클 계산 실패: {str(e)}"
        }), 500

@app.route('/api/v2/credit-cycle', methods=['GET'])
def get_credit_cycle():
    """
    신용/유동성 사이클 계산 및 국면 판별

    Returns:
        {
            "status": "success",
            "data": {
                "score": 0-100,
                "phase": "신용 경색|정상화|신용 과잉",
                "phase_en": "Credit Crunch|Normalizing|Credit Excess",
                "color": "red|amber|green",
                "description": "현재 국면 설명",
                "action": "투자 행동 추천",
                "confidence": 0-100,
                "indicators": {...},
                "last_updated": "ISO timestamp"
            }
        }
    """
    try:
        # CreditCycleService 초기화
        cycle_service = CreditCycleService(db_service)

        # 사이클 계산
        result = cycle_service.calculate_cycle()

        if result.get('score') is None:
            return jsonify({
                "status": "error",
                "message": result.get('description', '데이터를 불러올 수 없습니다')
            }), 503

        return jsonify({
            "status": "success",
            "data": result
        })

    except Exception as e:
        import traceback
        print(f"Error in get_credit_cycle: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"신용 사이클 계산 실패: {str(e)}"
        }), 500

# ===== 심리/밸류에이션 사이클 API =====
@app.route('/api/v2/sentiment-cycle', methods=['GET'])
def get_sentiment_cycle():
    """
    심리/밸류에이션 사이클 정보 반환 (VIX 기반)

    Returns:
        JSON: {
            status: "success",
            data: {
                score: 0-100,
                phase: "극단적 공포|중립|극단적 탐욕",
                phase_en: "Extreme Fear|Neutral|Extreme Greed",
                color: "green|amber|red",
                description: str,
                action: str,
                confidence: 0-100,
                indicators: {vix: 0-100},
                raw_data: {...},
                last_updated: ISO timestamp
            }
        }
    """
    try:
        from services.sentiment_cycle_service import SentimentCycleService

        cycle_service = SentimentCycleService(db_service)
        result = cycle_service.calculate_cycle()

        # 에러 응답인 경우
        if result.get('score') is None:
            return jsonify({
                "status": "error",
                "message": result.get('description')
            }), 503

        return jsonify({
            "status": "success",
            "data": result
        })

    except Exception as e:
        import traceback
        print(f"Error in get_sentiment_cycle: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"심리 사이클 계산 실패: {str(e)}"
        }), 500

# ========================================
# V3 API: Master Market Cycle System
# ========================================

@app.route('/api/v3/cycles/master', methods=['GET'])
def get_master_market_cycle():
    """
    Master Market Cycle 점수 조회 (Phase 1 임시 버전)

    Returns:
        {
            "status": "success",
            "data": {
                "mmc_score": 64.2,
                "phase": "확장기 (포지션 유지)",
                "macro": {...},
                "credit": {...},
                "sentiment": {"score": 50, "note": "Phase 2 예정"},
                "recommendation": "중립 포지션 유지",
                "updated_at": "2025-12-05T10:30:00"
            }
        }
    """
    try:
        from services.cycle_engine import calculate_master_cycle_v3

        result = calculate_master_cycle_v3(db_service)

        if 'error' in result:
            return jsonify({
                "status": "error",
                "message": result['error']
            }), 500

        return jsonify({
            "status": "success",
            "data": result
        })

    except Exception as e:
        import traceback
        print(f"Error in get_master_market_cycle: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"Master cycle 계산 실패: {str(e)}"
        }), 500


@app.route('/api/v4/master-cycle', methods=['GET'])
def get_master_market_cycle_v4():
    """
    v4: Master Market Cycle with Full Enhancements

    Phase 4+5 완성:
    - Macro: 실질금리 + 역전 지속기간
    - Credit: 스프레드 변화 속도 + 급변 탐지

    Returns:
        v3 구조 + Credit 강화 필드 (hy_velocity, ig_velocity, rapid_change)
    """
    try:
        from services.cycle_engine import calculate_master_cycle_v4

        result = calculate_master_cycle_v4(db_service)

        if 'error' in result:
            return jsonify({
                "status": "error",
                "message": result['error']
            }), 500

        return jsonify({
            "status": "success",
            "data": result
        })

    except Exception as e:
        import traceback
        print(f"Error in get_master_market_cycle_v4: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"Master cycle v4 계산 실패: {str(e)}"
        }), 500


@app.route('/api/v3/cycles/macro', methods=['GET'])
def get_macro_cycle_v3():
    """
    거시경제 사이클 점수 조회

    Returns:
        {
            "status": "success",
            "data": {
                "score": 72.5,
                "phase": "확장기",
                "signals": ["ISM 제조업 52.8", "실업률 3.8%"],
                "indicators": {...}
            }
        }
    """
    try:
        from services.cycle_engine import calculate_macro_score

        result = calculate_macro_score(db_service)

        return jsonify({
            "status": "success",
            "data": result
        })

    except Exception as e:
        import traceback
        print(f"Error in get_macro_cycle_v3: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"Macro cycle 계산 실패: {str(e)}"
        }), 500


@app.route('/api/v3/cycles/credit', methods=['GET'])
def get_credit_cycle_v3():
    """
    신용/유동성 사이클 점수 조회

    Returns:
        {
            "status": "success",
            "data": {
                "score": 58.3,
                "state": "중립",
                "signals": ["HY 350bp", "FCI -0.2"],
                "indicators": {...}
            }
        }
    """
    try:
        from services.cycle_engine import calculate_credit_score

        result = calculate_credit_score(db_service)

        return jsonify({
            "status": "success",
            "data": result
        })

    except Exception as e:
        import traceback
        print(f"Error in get_credit_cycle_v3: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"Credit cycle 계산 실패: {str(e)}"
        }), 500


# ===========================
# 자산 개별분석 API
# ===========================

@app.route('/api/asset-analysis', methods=['GET'])
def get_asset_analysis():
    """
    자산 개별분석 데이터 조회 (5개 탭 구조)

    Query Parameters:
        - asset_id: 자산 ID (필수)
        - user_id: 사용자 ID (필수)

    Returns:
        {
            "status": "success",
            "data": {
                "thesis": {...},      // 투자 가설
                "validation": {...},  // 검증: 펀더멘털
                "pricing": {...},     // 가격과 기대치
                "timing": {...},      // 타이밍 & 리스크
                "decision": {...}     // 결정 & 관리
            }
        }
    """
    try:
        asset_id = request.args.get('asset_id')
        user_id = request.args.get('user_id')

        if not asset_id or not user_id:
            return jsonify({
                "status": "error",
                "message": "asset_id와 user_id가 필요합니다"
            }), 400

        # 데이터베이스에서 분석 데이터 조회
        with db_service.get_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT thesis, validation, pricing, timing, decision, updated_at
                    FROM asset_analysis
                    WHERE asset_id = %s AND user_id = %s
                """, (asset_id, user_id))

                result = cur.fetchone()

                if result:
                    return jsonify({
                        "status": "success",
                        "data": {
                            "thesis": result['thesis'] or {},
                            "validation": result['validation'] or {},
                            "pricing": result['pricing'] or {},
                            "timing": result['timing'] or {},
                            "decision": result['decision'] or {},
                            "updated_at": result['updated_at'].isoformat() if result['updated_at'] else None
                        }
                    })
                else:
                    # 데이터가 없으면 빈 객체 반환
                    return jsonify({
                        "status": "success",
                        "data": {
                            "thesis": {},
                            "validation": {},
                            "pricing": {},
                            "timing": {},
                            "decision": {},
                            "updated_at": None
                        }
                    })

    except Exception as e:
        import traceback
        print(f"Error in get_asset_analysis: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"분석 데이터 조회 실패: {str(e)}"
        }), 500


@app.route('/api/asset-analysis', methods=['POST'])
def save_asset_analysis():
    """
    자산 개별분석 데이터 저장/업데이트 (5개 탭 구조)

    Request Body:
        {
            "asset_id": 123,
            "user_id": 1,
            "thesis": {...},      // 투자 가설
            "validation": {...},  // 검증: 펀더멘털
            "pricing": {...},     // 가격과 기대치
            "timing": {...},      // 타이밍 & 리스크
            "decision": {...}     // 결정 & 관리
        }

    Returns:
        {
            "status": "success",
            "message": "분석 데이터가 저장되었습니다"
        }
    """
    try:
        data = request.get_json()

        asset_id = data.get('asset_id')
        user_id = data.get('user_id')
        thesis = data.get('thesis', {})
        validation = data.get('validation', {})
        pricing = data.get('pricing', {})
        timing = data.get('timing', {})
        decision = data.get('decision', {})

        if not asset_id or not user_id:
            return jsonify({
                "status": "error",
                "message": "asset_id와 user_id가 필요합니다"
            }), 400

        # 데이터베이스에 저장 (UPSERT)
        with db_service.get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO asset_analysis (asset_id, user_id, thesis, validation, pricing, timing, decision, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
                    ON CONFLICT (asset_id, user_id)
                    DO UPDATE SET
                        thesis = EXCLUDED.thesis,
                        validation = EXCLUDED.validation,
                        pricing = EXCLUDED.pricing,
                        timing = EXCLUDED.timing,
                        decision = EXCLUDED.decision,
                        updated_at = NOW()
                """, (asset_id, user_id, json.dumps(thesis), json.dumps(validation), json.dumps(pricing), json.dumps(timing), json.dumps(decision)))

                conn.commit()

        return jsonify({
            "status": "success",
            "message": "분석 데이터가 저장되었습니다"
        })

    except Exception as e:
        import traceback
        print(f"Error in save_asset_analysis: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"분석 데이터 저장 실패: {str(e)}"
        }), 500


# ===========================
# Market Snapshot API (Finnhub)
# ===========================

def get_finnhub_key():
    return os.getenv("FINNHUB_API_KEY")


MARKET_CACHE = {
    "quote": {},
    "candles": {}
}


def get_cached(cache_key, ttl):
    entry = MARKET_CACHE.get(cache_key, {})
    now = time.time()
    for key, value in list(entry.items()):
        if now - value["ts"] > ttl:
            del entry[key]
    return entry


def set_cached(cache_key, key, data):
    MARKET_CACHE.setdefault(cache_key, {})[key] = {"ts": time.time(), "data": data}


def fetch_finnhub(endpoint, params):
    try:
        response = requests.get(endpoint, params=params, timeout=10)
        status = response.status_code
        try:
            payload = response.json()
        except ValueError:
            payload = {"raw": response.text[:500]}
        return status, payload
    except Exception as e:
        return 0, {"error": str(e)}


def fetch_stooq_candles(symbol, from_ts, to_ts):
    try:
        normalized = symbol.lower()
        if '.' not in normalized:
            normalized = f"{normalized}.us"
        urls = [
            "https://stooq.com/q/d/l/",
            "https://stooq.pl/q/d/l/",
        ]
        lines = []
        headers = {
            "User-Agent": "Mozilla/5.0 (compatible; investment-app/1.0)"
        }
        last_error = None
        for url in urls:
            try:
                response = requests.get(
                    url,
                    params={"s": normalized, "i": "d"},
                    headers=headers,
                    timeout=10
                )
                response.raise_for_status()
                lines = response.text.strip().splitlines()
                if lines:
                    break
            except Exception as e:
                last_error = str(e)
                continue

        if not lines:
            return {"s": "error", "error": last_error or "stooq_unreachable"}
        if len(lines) <= 1:
            return {"s": "no_data", "error": "stooq_empty"}

        reader = csv.DictReader(lines)
        all_rows = []
        for row in reader:
            date_str = row.get("Date")
            close_str = row.get("Close")
            if not date_str or not close_str:
                continue
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                ts = int(dt.replace(tzinfo=None).timestamp())
                all_rows.append((ts, float(close_str)))
            except Exception:
                continue

        if not all_rows:
            return {"s": "no_data", "error": "stooq_no_rows"}

        filtered = [
            (ts, close)
            for ts, close in all_rows
            if ts >= int(from_ts) and ts <= int(to_ts)
        ]
        if not filtered:
            filtered = all_rows[-180:]

        timestamps = [ts for ts, _ in filtered]
        closes = [close for _, close in filtered]

        return {"s": "ok", "t": timestamps, "c": closes}
    except Exception as e:
        return {"s": "error", "error": str(e)}


def fetch_yahoo_candles(symbol, from_ts, to_ts):
    try:
        normalized = symbol.upper()
        url = f"https://query2.finance.yahoo.com/v8/finance/chart/{normalized}"
        headers = {
            "User-Agent": "Mozilla/5.0 (compatible; investment-app/1.0)"
        }
        response = requests.get(
            url,
            params={
                "interval": "1d",
                "period1": from_ts,
                "period2": to_ts,
                "events": "div,splits"
            },
            headers=headers,
            timeout=10
        )
        response.raise_for_status()
        payload = response.json()
        result = (payload.get("chart") or {}).get("result") or []
        if not result:
            return {"s": "no_data", "error": "yahoo_no_result"}

        series = result[0]
        timestamps = series.get("timestamp") or []
        quotes = (series.get("indicators") or {}).get("quote") or []
        close_values = quotes[0].get("close") if quotes else []
        volume_values = quotes[0].get("volume") if quotes else []

        if not timestamps or not close_values:
            return {"s": "no_data", "error": "yahoo_no_prices"}

        filtered_t = []
        filtered_c = []
        filtered_v = []
        for idx, (ts, close) in enumerate(zip(timestamps, close_values)):
            if close is None:
                continue
            volume = None
            if volume_values and idx < len(volume_values):
                volume = volume_values[idx]
            filtered_t.append(ts)
            filtered_c.append(float(close))
            filtered_v.append(volume)

        if not filtered_t:
            return {"s": "no_data", "error": "yahoo_no_filtered"}

        if not any((volume or 0) > 0 for volume in filtered_v):
            quote = fetch_yahoo_quote(symbol)
            fallback_volume = (
                quote.get("regularMarketVolume")
                or quote.get("averageDailyVolume10Day")
                or quote.get("averageDailyVolume3Month")
            )
            if isinstance(fallback_volume, (int, float)) and fallback_volume > 0:
                padded_v = [0] * len(filtered_t)
                padded_v[-1] = fallback_volume
                filtered_v = padded_v

        return {"s": "ok", "t": filtered_t, "c": filtered_c, "v": filtered_v}
    except Exception as e:
        return {"s": "error", "error": str(e)}


def fetch_yahoo_quote(symbol):
    try:
        normalized = symbol.upper()
        url = "https://query2.finance.yahoo.com/v7/finance/quote"
        headers = {
            "User-Agent": "Mozilla/5.0 (compatible; investment-app/1.0)"
        }
        response = requests.get(
            url,
            params={"symbols": normalized},
            headers=headers,
            timeout=10
        )
        response.raise_for_status()
        payload = response.json()
        result = (payload.get("quoteResponse") or {}).get("result") or []
        if not result:
            return {}
        quote = result[0]
        return {
            "marketCap": quote.get("marketCap"),
            "fiftyTwoWeekHigh": quote.get("fiftyTwoWeekHigh"),
            "fiftyTwoWeekLow": quote.get("fiftyTwoWeekLow"),
            "regularMarketVolume": quote.get("regularMarketVolume"),
            "averageDailyVolume10Day": quote.get("averageDailyVolume10Day"),
            "averageDailyVolume3Month": quote.get("averageDailyVolume3Month"),
            "currency": quote.get("currency"),
        }
    except Exception:
        return {}


@app.route('/api/market/quote', methods=['GET'])
def get_market_quote():
    try:
        symbol = request.args.get('symbol')
        if not symbol:
            return jsonify({"status": "error", "message": "symbol이 필요합니다"}), 400

        api_key = get_finnhub_key()
        if not api_key:
            return jsonify({"status": "error", "message": "FINNHUB_API_KEY가 설정되지 않았습니다"}), 500

        cache = get_cached("quote", 30)
        if symbol in cache:
            cached = cache[symbol]["data"]
            if isinstance(cached, dict) and "quote" in cached:
                return jsonify({
                    "status": "success",
                    "data": cached.get("quote"),
                    "extra": cached.get("extra", {}),
                    "cached": True
                })
            return jsonify({"status": "success", "data": cached, "cached": True})

        status, data = fetch_finnhub(
            "https://finnhub.io/api/v1/quote",
            {"symbol": symbol, "token": api_key}
        )
        if status != 200:
            print(f"Finnhub quote error ({status}): {data}")
            return jsonify({"status": "error", "message": "시세 조회 실패", "detail": data}), status

        extra = fetch_yahoo_quote(symbol)
        set_cached("quote", symbol, {"quote": data, "extra": extra})
        return jsonify({"status": "success", "data": data, "extra": extra})
    except Exception as e:
        import traceback
        print(f"Error in get_market_quote: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"시세 조회 실패: {str(e)}"
        }), 500


@app.route('/api/market/candles', methods=['GET'])
def get_market_candles():
    try:
        symbol = request.args.get('symbol')
        resolution = request.args.get('resolution', 'D')
        from_ts = request.args.get('from')
        to_ts = request.args.get('to')
        source = request.args.get('source')

        if not symbol:
            return jsonify({"status": "error", "message": "symbol이 필요합니다"}), 400

        api_key = get_finnhub_key()
        if not api_key:
            return jsonify({"status": "error", "message": "FINNHUB_API_KEY가 설정되지 않았습니다"}), 500

        if not from_ts or not to_ts:
            return jsonify({"status": "error", "message": "from/to 파라미터가 필요합니다"}), 400

        cache_key = f"{symbol}:{resolution}:{from_ts}:{to_ts}:{source or 'auto'}"
        cache = get_cached("candles", 600)
        if cache_key in cache:
            return jsonify({"status": "success", "data": cache[cache_key]["data"], "cached": True})

        if source == "yahoo":
            yahoo = fetch_yahoo_candles(symbol, from_ts, to_ts)
            set_cached("candles", cache_key, yahoo)
            return jsonify({"status": "success", "data": yahoo, "source": "yahoo"})

        status, data = fetch_finnhub(
            "https://finnhub.io/api/v1/stock/candle",
            {
                "symbol": symbol,
                "resolution": resolution,
                "from": from_ts,
                "to": to_ts,
                "token": api_key,
            }
        )
        if status != 200:
            print(f"Finnhub candles error ({status}): {data}")
            if status == 403:
                fallback = fetch_yahoo_candles(symbol, from_ts, to_ts)
                return jsonify({"status": "success", "data": fallback, "source": "yahoo"})
            return jsonify({"status": "error", "message": "차트 데이터 조회 실패", "detail": data}), status

        if isinstance(data, dict) and data.get("s") != "ok":
            if data.get("s") == "no_data":
                fallback = fetch_yahoo_candles(symbol, from_ts, to_ts)
                return jsonify({"status": "success", "data": fallback, "source": "yahoo"})
            return jsonify({"status": "success", "data": data})

        set_cached("candles", cache_key, data)
        return jsonify({"status": "success", "data": data})
    except Exception as e:
        import traceback
        print(f"Error in get_market_candles: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"차트 데이터 조회 실패: {str(e)}"
        }), 500


# =====================================
# Industry Analysis API (산업군 분석)
# =====================================

@app.route('/api/industry-analysis', methods=['GET'])
def get_industry_analysis():
    """특정 산업 분석 조회"""
    try:
        user_id = request.args.get('user_id', type=int)
        major_category = request.args.get('major_category')
        sub_industry = request.args.get('sub_industry')

        if not all([user_id, major_category, sub_industry]):
            return jsonify({
                "status": "error",
                "message": "user_id, major_category, sub_industry 파라미터가 필요합니다"
            }), 400

        result = db_service.get_industry_analysis(user_id, major_category, sub_industry)

        if result:
            return jsonify({
                "status": "success",
                "data": result
            })
        else:
            # 데이터 없으면 빈 템플릿 반환
            return jsonify({
                "status": "success",
                "data": {
                    "analysis_data": {
                        "core_technology": {"definition": "", "stage": "상용화", "innovation_path": ""},
                        "macro_impact": {"interest_rate": "", "exchange_rate": "", "commodities": "", "policy": ""},
                        "growth_drivers": {"internal": "", "external": "", "kpi": ""},
                        "value_chain": {"flow": "", "profit_pool": "", "bottleneck": ""},
                        "supply_demand": {
                            "demand": {"end_user": "", "long_term": "", "sensitivity": ""},
                            "supply": {"players": "", "capacity": "", "barriers": ""},
                            "catalysts": ""
                        },
                        "market_map": {"structure": "", "competition": "", "moat": "", "lifecycle": ""}
                    },
                    "leading_stocks": [],
                    "emerging_stocks": []
                }
            })

    except Exception as e:
        import traceback
        print(f"Error in get_industry_analysis: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"산업 분석 조회 실패: {str(e)}"
        }), 500


@app.route('/api/industry-analysis', methods=['POST'])
def save_industry_analysis():
    """산업 분석 저장 (UPSERT)"""
    try:
        data = request.get_json()

        user_id = data.get('user_id')
        major_category = data.get('major_category')
        sub_industry = data.get('sub_industry')
        analysis_data = data.get('analysis_data', {})
        leading_stocks = data.get('leading_stocks', [])
        emerging_stocks = data.get('emerging_stocks', [])

        if not all([user_id, major_category, sub_industry]):
            return jsonify({
                "status": "error",
                "message": "user_id, major_category, sub_industry는 필수입니다"
            }), 400

        success = db_service.save_industry_analysis(
            user_id,
            major_category,
            sub_industry,
            analysis_data,
            leading_stocks,
            emerging_stocks
        )

        if success:
            return jsonify({
                "status": "success",
                "message": "산업 분석이 저장되었습니다"
            })
        else:
            return jsonify({
                "status": "error",
                "message": "산업 분석 저장 실패"
            }), 500

    except Exception as e:
        import traceback
        print(f"Error in save_industry_analysis: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"산업 분석 저장 실패: {str(e)}"
        }), 500


@app.route('/api/industry-categories', methods=['GET'])
def get_industry_categories():
    """특정 산업군의 모든 하위 산업 목록 조회"""
    try:
        user_id = request.args.get('user_id', type=int)
        major_category = request.args.get('major_category')

        if not all([user_id, major_category]):
            return jsonify({
                "status": "error",
                "message": "user_id, major_category 파라미터가 필요합니다"
            }), 400

        results = db_service.get_all_industries_by_major(user_id, major_category)

        return jsonify({
            "status": "success",
            "data": results
        })

    except Exception as e:
        import traceback
        print(f"Error in get_industry_categories: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"산업군 목록 조회 실패: {str(e)}"
        }), 500


if __name__ == '__main__':
    # Render 등 PaaS 환경에서 주어지는 동적 포트를 우선 사용
    port = int(os.environ.get("PORT", 5001))
    app.run(debug=True, host='0.0.0.0', port=port)
